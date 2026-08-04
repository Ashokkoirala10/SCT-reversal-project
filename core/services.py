"""
Core reconciliation logic.

Turns an "ibft-transaction_*.xlsx" export into a "need_to_reversal_*.xlsx"
workbook containing:

  - failed      : FAILED transactions, excluding "Insufficient funds"
  - coop        : manual-reversal transactions for every aggregator other
                  than IME REMIT / CITY REMIT, grouped by Member Name
  - imeremit    : manual-reversal transactions for the IME REMIT aggregator
  - cityremit   : manual-reversal transactions for the CITY REMIT aggregator
  - timeout     : TIMEOUT transactions, unchanged

This logic was reverse-engineered from a real ibft-transaction file and its
matching need_to_reversal output and validated row-for-row (104/104 reversal
rows, 29/29 failed rows, 1/1 timeout row) before being written here.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# The house/clearing account that sits on the debit side of a reversal
# entry when the transaction's Debtor Bank is (Global IME) Bank — this is
# the normal/current case.
GLOBAL_BANK_DEBIT_ACCOUNT = "0002335524115"
GLOBAL_BANK_KEYWORD = "GLOBAL"
# kept as an alias since other code/tests refer to this name
CLEARING_DEBIT_ACCOUNT = GLOBAL_BANK_DEBIT_ACCOUNT

# Prabhu Bank is a rarer case (it was the Debtor Bank before Global IME
# Bank took over that role) — when it shows up, the reversal's Debit
# Account Number must be this dedicated account instead.
PRABHU_BANK_DEBIT_ACCOUNT = "99901170130555"
PRABHU_BANK_KEYWORD = "PRABHU"

# Network types for which the charge amount is always forced to 0 on the
# reversal, regardless of what the original file shows.
ZERO_CHARGE_NETWORK_KEYWORDS = ("NCHL",)

SOURCE_SHEET_NAME_CANDIDATES = ["Transactions", "transactions"]

# Column names expected in the source workbook (must match, case sensitive,
# the header row of the "Transactions" sheet).
REQUIRED_COLUMNS = [
    "S NO",
    "Transaction Date",
    "Member Name",
    "Aggregator",
    "Member Transaction Id",
    "Network Reference Id",
    "Session Id",
    "Payment Processor",
    "Transaction Amount",
    "Charge Amount",
    "Debtor Bank",
    "Debit Status",
    "Debit Response Code",
    "Debit Account Number",
    "Debitor Account Name",
    "Creditor Bank",
    "Credit Status",
    "Credit Response Code",
    "Credit Account Number",
    "Creditor Account Name",
    "Source Message",
    "Destination Message",
    "Overall Status",
]

# Columns pulled, in order, from the source row for the reversal-style sheets
# (coop / imeremit / cityremit).
REVERSAL_SOURCE_COLUMNS = [
    "S NO",
    "Transaction Date",
    "Member Name",
    "Aggregator",
    "Member Transaction Id",
    "Network Reference Id",
    "Session Id",
    "Payment Processor",
    "Transaction Amount",
    "Charge Amount",
    "Debtor Bank",
    "Creditor Bank",
    "Debit Account Number",
    "Credit Account Number",
]

# Header row actually written for the reversal-style sheets. The last three
# columns (Debit Account Number, Credit Account Number, Narration) are
# highlighted and are not a straight passthrough of the source row.
REVERSAL_SHEET_HEADERS = [
    "S NO",
    "Transaction Date",
    "Member Name",
    "Aggregator",
    "Member Transaction Id",
    "Network Reference Id",
    "Session Id",
    "Payment Processor",
    "Transaction Amount",
    "Charge Amount",
    "Debtor Bank",
    "Creditor Bank",
    "Debit Account Number",
    "Credit Account Number",
    "Narration",
]

# How many trailing columns get the yellow header highlight.
HIGHLIGHTED_TRAILING_COLUMNS = 3

IME_AGGREGATOR = "IME REMIT"
CITY_AGGREGATOR = "CITY REMIT"

# Sheet holding manual-reversal rows whose Debtor Bank is Prabhu Bank —
# kept separate from coop/imeremit/cityremit so Prabhu reversals can be
# worked (and bank-statement-checked) on their own.
PRABHU_SHEET_NAME = "prabhu"

# Sheet holding SYSTEM-initiated REVERSAL rows (the switch already
# reversed these automatically — see is_manual_reversal()) that are
# either On-Us (Global-to-Global or Prabhu-to-Prabhu, see is_on_us()) or
# routed over the NCHL network (see is_zero_charge_network()). These were
# previously only counted (reversal_system_count/amount/charge) and never
# written anywhere; kept on their own sheet so they can be reviewed the
# same way the manual-reversal sheets are. Max 31 chars (Excel limit).
ONUS_SYSTEM_REVERSAL_SHEET_NAME = "Onus Checked-System Reversal"


class ProcessingError(Exception):
    """Raised when the uploaded workbook doesn't look like a valid export."""


@dataclass
class ProcessingStats:
    total_rows: int = 0
    # The latest "Transaction Date" value found anywhere in this upload's
    # source file — i.e. how far into the day this export actually
    # reaches (e.g. "downloaded at 10am, so data through 09:58am"), shown
    # on the dashboard's day-by-day breakdown. Not the same as wall-clock
    # processing time.
    data_through_at: Any = None
    success_count: int = 0
    success_amount: float = 0.0
    failed_total: int = 0
    failed_insufficient_funds: int = 0
    failed_kept: int = 0
    failed_amount: float = 0.0
    failed_charge: float = 0.0
    # Amount/charge summed only across the rows actually kept in the
    # "failed" sheet (i.e. failed_total minus insufficient-funds) — needed
    # so the bank-statement check can subtract out "failed but credited"
    # rows precisely.
    failed_kept_amount: float = 0.0
    failed_kept_charge: float = 0.0
    reversal_total: int = 0
    reversal_manual_kept: int = 0
    reversal_manual_amount: float = 0.0
    reversal_manual_charge: float = 0.0
    reversal_system_count: int = 0
    reversal_system_amount: float = 0.0
    reversal_system_charge: float = 0.0
    # Subset of reversal_system_* above that is either On-Us (Global-to-
    # Global or Prabhu-to-Prabhu) or NCHL-routed — written out to their own
    # sheet (ONUS_SYSTEM_REVERSAL_SHEET_NAME) in the generated file so they
    # can be reviewed, instead of only being counted.
    reversal_system_onus_count: int = 0
    reversal_system_onus_amount: float = 0.0
    reversal_system_onus_charge: float = 0.0
    coop_count: int = 0
    imeremit_count: int = 0
    cityremit_count: int = 0
    timeout_count: int = 0
    timeout_amount: float = 0.0
    coop_member_count: int = 0
    prabhu_rerouted: int = 0
    # Manual-reversal rows whose Debtor Bank is Prabhu Bank — written to
    # their own "prabhu" sheet instead of coop/imeremit/cityremit.
    prabhu_reversal_count: int = 0
    prabhu_reversal_amount: float = 0.0
    duplicate_skipped: int = 0
    # Rows skipped because their Network Reference Id was already seen in
    # an *earlier upload's source file* (any status) — this catches the
    # "downloaded an overlapping time window" case, as opposed to
    # duplicate_skipped above, which only catches manual-reversal rows
    # repeated across generated reversal files.
    duplicate_source_skipped: int = 0
    unrecognized_debtor_bank_rows: int = 0
    unrecognized_debtor_banks: list[str] | None = None
    failed_reason_breakdown: dict[str, int] | None = None
    failed_onus_count: int = 0
    failed_onus_amount: float = 0.0
    failed_offus_count: int = 0
    failed_offus_amount: float = 0.0
    failed_reason_breakdown_onus: dict[str, int] | None = None
    failed_reason_breakdown_offus: dict[str, int] | None = None
    # Not persisted to the DB directly (see views.py) — every normalized
    # Network Reference Id that was actually kept/counted this run, so the
    # caller can remember them for next time's overlapping-window dedup.
    kept_reference_ids: set[str] = field(default_factory=set)
    # Not persisted via as_dict() either — per (Member Name, Aggregator)
    # breakdown of success/failed/manual-reversal count+amount, keyed by
    # "<member>\x1f<aggregator>". The view turns this into MemberAggregatorStat
    # rows (see core/models.py) for the dashboard's Member & Aggregator report.
    member_aggregator_breakdown: dict[str, dict] = field(default_factory=dict)

    def as_dict(self) -> dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "data_through_at": self.data_through_at,
            "success_count": self.success_count,
            "success_amount": self.success_amount,
            "failed_total": self.failed_total,
            "failed_insufficient_funds": self.failed_insufficient_funds,
            "failed_kept": self.failed_kept,
            "failed_amount": self.failed_amount,
            "failed_charge": self.failed_charge,
            "failed_kept_amount": self.failed_kept_amount,
            "failed_kept_charge": self.failed_kept_charge,
            "reversal_total": self.reversal_total,
            "reversal_manual_kept": self.reversal_manual_kept,
            "reversal_manual_amount": self.reversal_manual_amount,
            "reversal_manual_charge": self.reversal_manual_charge,
            "reversal_system_count": self.reversal_system_count,
            "reversal_system_amount": self.reversal_system_amount,
            "reversal_system_charge": self.reversal_system_charge,
            "reversal_system_onus_count": self.reversal_system_onus_count,
            "reversal_system_onus_amount": self.reversal_system_onus_amount,
            "reversal_system_onus_charge": self.reversal_system_onus_charge,
            "coop_count": self.coop_count,
            "imeremit_count": self.imeremit_count,
            "cityremit_count": self.cityremit_count,
            "timeout_count": self.timeout_count,
            "timeout_amount": self.timeout_amount,
            "coop_member_count": self.coop_member_count,
            "prabhu_rerouted": self.prabhu_rerouted,
            "prabhu_reversal_count": self.prabhu_reversal_count,
            "prabhu_reversal_amount": self.prabhu_reversal_amount,
            "duplicate_skipped": self.duplicate_skipped,
            "duplicate_source_skipped": self.duplicate_source_skipped,
            "unrecognized_debtor_bank_rows": self.unrecognized_debtor_bank_rows,
            "unrecognized_debtor_banks": self.unrecognized_debtor_banks or [],
            "failed_reason_breakdown": self.failed_reason_breakdown or {},
            "failed_onus_count": self.failed_onus_count,
            "failed_onus_amount": self.failed_onus_amount,
            "failed_offus_count": self.failed_offus_count,
            "failed_offus_amount": self.failed_offus_amount,
            "failed_reason_breakdown_onus": self.failed_reason_breakdown_onus or {},
            "failed_reason_breakdown_offus": self.failed_reason_breakdown_offus or {},
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def transform_member_id(raw_id: Any) -> str:
    """Strip the leading "noise" off a member transaction id.

    Rule (derived from real data, validated on 104 real rows):
      1. Strip any leading zeros.
      2. Strip everything from the front up to (but not including) the
         first digit that remains.

    Examples:
        'REQ_1783249894525'                 -> '1783249894525'
        'IME-00113275637-572'               -> '00113275637-572'
        'TXN_2421483084574022'              -> '2421483084574022'
        'ACQ_SETTL-2607050005472234LFE'     -> '2607050005472234LFE'
        '0000000000CEMTP402131874/7028457327' -> '402131874/7028457327'
        '26180083314147965' (plain numeric) -> '26180083314147965'
    """
    s = "" if raw_id is None else str(raw_id).strip()

    i = 0
    while i < len(s) and s[i] == "0":
        i += 1
    s = s[i:]

    j = 0
    while j < len(s) and not s[j].isdigit():
        j += 1
    return s[j:]


def build_narration(member_transaction_id: Any, session_id: Any) -> str:
    return f"REV{transform_member_id(member_transaction_id)}-{session_id}"


def is_insufficient_funds(source_message: Any) -> bool:
    return "insufficient funds" in str(source_message or "").lower()


def is_manual_reversal(source_message: Any) -> bool:
    return "manual reversal" in str(source_message or "").lower()


def is_zero_charge_network(payment_processor: Any) -> bool:
    """NCHL-routed transactions always carry a 0 charge on the reversal."""
    value = str(payment_processor or "").upper()
    return any(keyword in value for keyword in ZERO_CHARGE_NETWORK_KEYWORDS)


def resolve_charge_amount(row, col) -> Any:
    charge = row[col["Charge Amount"]]
    if is_zero_charge_network(row[col["Payment Processor"]]):
        return 0
    return charge


def to_float(value: Any) -> float:
    """Safely coerce a cell value to float for summing (amount/charge
    columns). Blank/None/non-numeric values count as 0 rather than
    raising, so a stray text cell can't blow up the whole run."""
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_transaction_datetime(value: Any):
    """Parse a "Transaction Date" cell (e.g. "2026-07-17T13:59:05.461")
    into a datetime, or None if it isn't parseable. Used to figure out
    how far into the day an uploaded ibft export actually reaches — i.e.
    "processed up to" — rather than reporting wall-clock processing time."""
    import datetime as _dt

    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return _dt.datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return _dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Dynamic bank accounts (admin-configurable — see core.models.BankAccount)
# ---------------------------------------------------------------------------
#
# The two constants above (GLOBAL_BANK_*/PRABHU_BANK_*) remain the built-in
# fallback and are what's used until an admin actually configures anything
# in Admin > Core > Bank accounts, so behavior is unchanged out of the box.
# Once configured, resolve_debit_account() below prefers the DB rows —
# editing a row's Debit Account Number (or adding a brand-new bank) takes
# effect immediately without a code change or deploy, since the cache is
# cleared by core/models.py's BankAccount post_save/post_delete signal.

from functools import lru_cache


@lru_cache(maxsize=1)
def _bank_accounts_cached() -> tuple[tuple[str, str, bool], ...]:
    """(keyword, debit_account_number, is_own_bank) for every active
    BankAccount row, read from the DB once per process and cached until an
    admin edits/adds/deletes one."""
    from .models import BankAccount  # lazy: avoids a top-level services<->models dependency

    return tuple(
        (a.keyword.strip().upper(), (a.debit_account_number or "").strip(), bool(a.is_own_bank))
        for a in BankAccount.objects.filter(is_active=True).order_by("id")
        if a.keyword and a.debit_account_number
    )


def bank_accounts_cache_clear() -> None:
    """Called by core/models.py's BankAccount post_save/post_delete signal
    so admin edits to bank accounts take effect immediately."""
    _bank_accounts_cached.cache_clear()


def _bank_accounts() -> tuple[tuple[str, str, bool], ...]:
    try:
        accounts = _bank_accounts_cached()
    except Exception:
        # DB not migrated / not available yet (e.g. some management
        # commands run before `migrate`) — fall back to the hardcoded pair.
        accounts = ()
    if accounts:
        return accounts
    return (
        (GLOBAL_BANK_KEYWORD, GLOBAL_BANK_DEBIT_ACCOUNT, True),
        (PRABHU_BANK_KEYWORD, PRABHU_BANK_DEBIT_ACCOUNT, True),
    )


def is_prabhu_bank(debtor_bank: Any) -> bool:
    return PRABHU_BANK_KEYWORD in str(debtor_bank or "").upper()


def is_global_bank(debtor_bank: Any) -> bool:
    return GLOBAL_BANK_KEYWORD in str(debtor_bank or "").upper()


def resolve_debit_account(debtor_bank: Any) -> tuple[str, bool]:
    """Returns (debit_account_number, recognized).

    Checks every admin-configured Bank account (Admin > Core > Bank
    accounts) in order and returns the first whose Keyword matches the
    Debtor Bank — so editing an account number, or adding a brand-new
    bank, there takes effect immediately with no code change. Falls back
    to the original hardcoded Global IME Bank / Prabhu Bank pair if
    nothing has been configured in the admin yet, so out-of-the-box
    behavior is unchanged. Anything that matches neither still falls back
    to the standard (Global IME Bank) account (so processing isn't
    blocked), but `recognized=False` is bubbled up so it gets counted and
    surfaced on the audit log / dashboard for manual review.
    """
    name = str(debtor_bank or "").upper()
    fallback_account = GLOBAL_BANK_DEBIT_ACCOUNT
    for keyword, account, _is_own in _bank_accounts():
        if keyword == GLOBAL_BANK_KEYWORD:
            fallback_account = account
        if keyword and keyword in name:
            return account, True
    return fallback_account, False


def normalize_reference_id(value: Any) -> str:
    """Normalize a Network Reference Id (or similar) for reliable
    comparison — same value read as ' ABC123 ', 'abc123', or a numeric type
    should all compare equal."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().upper()


def normalize_failure_reason(source_message: Any) -> str:
    """Bucket a raw Source Message into one of a small, fixed set of
    reason labels for analytics, matched case-insensitively:

        Timeout, Insufficient fund, Card issuer timeout, Response timeout,
        Transaction amount exceeded, Other

    Everything that doesn't match a known phrase (regardless of case)
    falls into "Other" — so the dashboard breakdown always sums cleanly
    to the total failed count instead of fragmenting into many one-off
    raw-message buckets.
    """
    msg = str(source_message or "").strip()
    if not msg:
        return "Other"

    lowered = msg.lower()

    # Order matters: check the more specific phrases before the generic
    # "timeout" one so e.g. "card issuer timeout" doesn't just land in
    # the generic "Timeout" bucket.
    known_buckets = [
        ("insufficient fund", "Insufficient fund"),
        ("card issuer timeout", "Card issuer timeout"),
        ("issuer timeout", "Card issuer timeout"),
        ("response timeout", "Response timeout"),
        ("transaction amount exceed", "Transaction amount exceeded"),
        ("amount exceed", "Transaction amount exceeded"),
        ("time out", "Timeout"),
        ("timeout", "Timeout"),
    ]
    for needle, label in known_buckets:
        if needle in lowered:
            return label

    return "Other"


def is_on_us(debtor_bank: Any, creditor_bank: Any) -> bool:
    """On-Us = the Debtor Bank AND the Creditor Bank are the *same* one of
    our own banks — Global-to-Global or Prabhu-to-Prabhu by default, plus
    any other bank an admin has marked "Is own bank" in Admin > Core >
    Bank accounts (see core.models.BankAccount). Off-Us = Global-to-other
    / Prabhu-to-other (or any other combination). Used to split the
    failed-transaction breakdown by where the failure actually originated,
    and to decide which rows get the On-Us "already successful" (duplicate
    DR) verification elsewhere in this module."""
    if (is_global_bank(debtor_bank) and is_global_bank(creditor_bank)) or (
        is_prabhu_bank(debtor_bank) and is_prabhu_bank(creditor_bank)
    ):
        return True
    debtor_name = str(debtor_bank or "").upper()
    creditor_name = str(creditor_bank or "").upper()
    for keyword, _account, is_own in _bank_accounts():
        if is_own and keyword and keyword in debtor_name and keyword in creditor_name:
            return True
    return False


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Locate the header row (some exports have a merged 'Period: ...'
    banner row above the real header) and return (row_index, col_index_map).
    col_index_map maps column name -> 0-based index within that row.
    """
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True)
    ):
        names = [str(c).strip() if c is not None else "" for c in row]
        if "S NO" in names and "Overall Status" in names:
            col_map = {name: i for i, name in enumerate(names) if name}
            missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
            if missing:
                raise ProcessingError(
                    "The uploaded file is missing expected column(s): "
                    + ", ".join(missing)
                )
            return row_idx, col_map
    raise ProcessingError(
        "Could not find the header row (expected a row containing "
        "'S NO' and 'Overall Status'). Please upload the original "
        "ibft-transaction export."
    )


def _load_transactions(input_path: str | Path):
    wb = openpyxl.load_workbook(input_path, data_only=True)

    ws = None
    for name in SOURCE_SHEET_NAME_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        # fall back to the first sheet
        ws = wb[wb.sheetnames[0]]

    header_row_idx, col_map = _find_header_row(ws)

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if row[col_map["S NO"]] is None and all(v is None for v in row):
            continue
        rows.append(row)

    return rows, col_map


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def process_ibft_file(
    input_path: str | Path,
    output_path: str | Path,
    previously_reversed_ids: set[str] | None = None,
    seen_reference_ids: set[str] | None = None,
) -> ProcessingStats:
    """Read the ibft-transaction export at `input_path` and write the
    reconciled need_to_reversal workbook to `output_path`.

    `previously_reversed_ids` is the set of (normalized) Network Reference
    Ids that already appeared in the *previous* generated reversal file.
    Any manual-reversal row whose Network Reference Id is in that set is
    skipped here, so the same transaction can never be reversed twice just
    because it was still showing up as REVERSAL in a later export.

    `seen_reference_ids` is the set of (normalized) Network Reference Ids
    already processed in *any earlier upload* (any status — success,
    failed, reversal, timeout). Daily exports are typically downloaded as
    "since midnight through whenever I downloaded", so two exports on
    consecutive days overlap for the hours between midnight and the
    previous download time. Any row whose Network Reference Id is already
    in this set is skipped entirely (counted as `duplicate_source_skipped`)
    before it's categorized by Overall Status, so that overlap is never
    counted or reversed twice.
    """
    previously_reversed_ids = previously_reversed_ids or set()
    seen_reference_ids = seen_reference_ids or set()
    rows, col = _load_transactions(input_path)
    stats = ProcessingStats(total_rows=len(rows))

    if "Transaction Date" in col:
        date_idx = col["Transaction Date"]
        parsed_dates = (parse_transaction_datetime(r[date_idx]) for r in rows)
        latest = max((d for d in parsed_dates if d is not None), default=None)
        if latest is not None:
            try:
                from django.utils import timezone as _tz

                if _tz.is_naive(latest):
                    latest = _tz.make_aware(latest)
            except Exception:
                pass  # services.py stays usable outside Django too
        stats.data_through_at = latest

    def g(row, name):
        return row[col[name]]

    failed_kept_rows = []
    timeout_rows = []
    reversal_coop_rows = []
    reversal_ime_rows = []
    reversal_city_rows = []
    reversal_prabhu_rows = []
    reversal_system_onus_rows = []
    reason_breakdown: dict[str, int] = {}
    reason_breakdown_onus: dict[str, int] = {}
    reason_breakdown_offus: dict[str, int] = {}
    unrecognized_banks_seen: dict[str, int] = {}
    kept_reference_ids: set[str] = set()
    member_aggregator_breakdown: dict[str, dict] = {}

    def _magg_bucket(member_name, aggregator_name):
        key = f"{str(member_name or '').strip()}\x1f{str(aggregator_name or '').strip()}"
        return member_aggregator_breakdown.setdefault(
            key,
            {
                "success_count": 0, "success_amount": 0.0,
                "failed_count": 0, "failed_amount": 0.0,
                "reversal_count": 0, "reversal_amount": 0.0,
            },
        )

    for row in rows:
        row_ref_id = normalize_reference_id(g(row, "Network Reference Id"))
        if row_ref_id and row_ref_id in seen_reference_ids:
            stats.duplicate_source_skipped += 1
            continue
        if row_ref_id:
            kept_reference_ids.add(row_ref_id)

        overall = str(g(row, "Overall Status") or "").strip().upper()

        if overall == "SUCCESS":
            stats.success_count += 1
            amount = to_float(g(row, "Transaction Amount"))
            stats.success_amount += amount
            bucket = _magg_bucket(g(row, "Member Name"), g(row, "Aggregator"))
            bucket["success_count"] += 1
            bucket["success_amount"] += amount

        elif overall == "FAILED":
            stats.failed_total += 1
            failed_amount = to_float(g(row, "Transaction Amount"))
            stats.failed_amount += failed_amount
            stats.failed_charge += to_float(resolve_charge_amount(row, col))
            bucket = _magg_bucket(g(row, "Member Name"), g(row, "Aggregator"))
            bucket["failed_count"] += 1
            bucket["failed_amount"] += failed_amount
            reason = normalize_failure_reason(g(row, "Source Message"))
            reason_breakdown[reason] = reason_breakdown.get(reason, 0) + 1
            txn_amount = to_float(g(row, "Transaction Amount"))
            if is_on_us(g(row, "Debtor Bank"), g(row, "Creditor Bank")):
                stats.failed_onus_count += 1
                stats.failed_onus_amount += txn_amount
                reason_breakdown_onus[reason] = reason_breakdown_onus.get(reason, 0) + 1
            else:
                stats.failed_offus_count += 1
                stats.failed_offus_amount += txn_amount
                reason_breakdown_offus[reason] = reason_breakdown_offus.get(reason, 0) + 1
            if is_insufficient_funds(g(row, "Source Message")):
                stats.failed_insufficient_funds += 1
            else:
                failed_kept_rows.append(row)
                stats.failed_kept_amount += to_float(g(row, "Transaction Amount"))
                stats.failed_kept_charge += to_float(resolve_charge_amount(row, col))

        elif overall == "TIMEOUT":
            stats.timeout_count += 1
            stats.timeout_amount += to_float(g(row, "Transaction Amount"))
            timeout_rows.append(row)

        elif overall == "REVERSAL":
            stats.reversal_total += 1
            if is_manual_reversal(g(row, "Source Message")):
                ref_id = normalize_reference_id(g(row, "Network Reference Id"))
                if ref_id and ref_id in previously_reversed_ids:
                    stats.duplicate_skipped += 1
                    continue
                stats.reversal_manual_kept += 1
                reversal_amount = to_float(g(row, "Transaction Amount"))
                stats.reversal_manual_amount += reversal_amount
                stats.reversal_manual_charge += to_float(resolve_charge_amount(row, col))
                bucket = _magg_bucket(g(row, "Member Name"), g(row, "Aggregator"))
                bucket["reversal_count"] += 1
                bucket["reversal_amount"] += reversal_amount
                debtor_bank_name = g(row, "Debtor Bank")
                _, recognized = resolve_debit_account(debtor_bank_name)
                if is_prabhu_bank(debtor_bank_name):
                    stats.prabhu_rerouted += 1
                if not recognized:
                    label = str(debtor_bank_name or "(blank)").strip() or "(blank)"
                    unrecognized_banks_seen[label] = unrecognized_banks_seen.get(label, 0) + 1
                aggregator = str(g(row, "Aggregator") or "").strip().upper()
                if is_prabhu_bank(debtor_bank_name):
                    # Prabhu Bank reversals get their own sheet, regardless
                    # of aggregator, so they can be worked (and checked
                    # against a Prabhu bank statement) separately.
                    reversal_prabhu_rows.append(row)
                elif aggregator == IME_AGGREGATOR:
                    reversal_ime_rows.append(row)
                elif aggregator == CITY_AGGREGATOR:
                    reversal_city_rows.append(row)
                else:
                    reversal_coop_rows.append(row)
            else:
                # System-initiated reversal — already reversed automatically
                # by the switch, so it doesn't need a manual reversal row in
                # the output file, but it's still tracked here for daily
                # reconciliation (how much moved via system reversal vs.
                # manual reversal).
                stats.reversal_system_count += 1
                sys_amount = to_float(g(row, "Transaction Amount"))
                sys_charge = to_float(resolve_charge_amount(row, col))
                stats.reversal_system_amount += sys_amount
                stats.reversal_system_charge += sys_charge

                # On-Us (Global-to-Global or Prabhu-to-Prabhu) or NCHL-
                # routed system reversals get verified/kept on their own
                # sheet instead of only being counted, so they can be
                # reviewed the same way manual reversals are.
                is_onus = is_on_us(g(row, "Debtor Bank"), g(row, "Creditor Bank"))
                is_nchl = is_zero_charge_network(g(row, "Payment Processor"))
                if is_onus or is_nchl:
                    stats.reversal_system_onus_count += 1
                    stats.reversal_system_onus_amount += sys_amount
                    stats.reversal_system_onus_charge += sys_charge
                    reversal_system_onus_rows.append(row)

    stats.failed_kept = len(failed_kept_rows)
    stats.imeremit_count = len(reversal_ime_rows)
    stats.cityremit_count = len(reversal_city_rows)
    stats.coop_count = len(reversal_coop_rows)
    stats.prabhu_reversal_count = len(reversal_prabhu_rows)
    stats.prabhu_reversal_amount = sum(
        to_float(row[col["Transaction Amount"]]) for row in reversal_prabhu_rows
    )
    stats.unrecognized_debtor_bank_rows = sum(unrecognized_banks_seen.values())
    stats.unrecognized_debtor_banks = sorted(unrecognized_banks_seen.keys())
    stats.failed_reason_breakdown = reason_breakdown
    stats.failed_reason_breakdown_onus = reason_breakdown_onus
    stats.failed_reason_breakdown_offus = reason_breakdown_offus
    stats.kept_reference_ids = kept_reference_ids
    stats.member_aggregator_breakdown = member_aggregator_breakdown

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    _write_raw_sheet(out_wb, "failed", failed_kept_rows, col)
    _write_reversal_sheet_grouped_by_member(out_wb, "coop", reversal_coop_rows, col, stats)
    _write_reversal_sheet_flat(out_wb, "imeremit", reversal_ime_rows, col)
    _write_reversal_sheet_flat(out_wb, "cityremit", reversal_city_rows, col)
    _write_reversal_sheet_flat(out_wb, PRABHU_SHEET_NAME, reversal_prabhu_rows, col)
    _write_raw_sheet(out_wb, "timeout", timeout_rows, col)
    _write_raw_sheet(out_wb, ONUS_SYSTEM_REVERSAL_SHEET_NAME, reversal_system_onus_rows, col)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)

    return stats


# ---------------------------------------------------------------------------
# Double-reversal prevention: read Network Reference Ids out of a
# previously generated reversal workbook.
# ---------------------------------------------------------------------------

_REVERSAL_SHEET_NAMES = ("coop", "imeremit", "cityremit", PRABHU_SHEET_NAME)


def extract_reversal_network_reference_ids(reversal_file_path: str | Path) -> set[str]:
    """Open a previously generated need_to_reversal_*.xlsx and collect every
    Network Reference Id across its reversal-style sheets (coop / imeremit /
    cityremit), so the next run can exclude them and avoid a double
    reversal. Silently returns an empty set if the file can't be read
    (e.g. it's missing or was deleted) rather than blocking new processing.
    """
    ids: set[str] = set()
    try:
        wb = openpyxl.load_workbook(reversal_file_path, data_only=True, read_only=True)
    except Exception:
        return ids

    try:
        for sheet_name in _REVERSAL_SHEET_NAMES:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_seen = False
            ref_col_idx = None
            for row in ws.iter_rows(values_only=True):
                if all(v is None for v in row):
                    # blank separator row between coop member groups
                    header_seen = False
                    continue
                if not header_seen:
                    names = [str(c).strip() if c is not None else "" for c in row]
                    if "Network Reference Id" in names:
                        ref_col_idx = names.index("Network Reference Id")
                        header_seen = True
                    continue
                if ref_col_idx is not None and ref_col_idx < len(row):
                    ref_id = normalize_reference_id(row[ref_col_idx])
                    if ref_id:
                        ids.add(ref_id)
    finally:
        wb.close()

    return ids


# ---------------------------------------------------------------------------
# Bank statement cross-check
#
# The bank statement export (e.g. "global_BankStatement_<from>_<to>.csv")
# has columns: S.N, ENTRY TYPE (CR/DR), REMARKS, AMOUNT, DATE. A row's
# Network Reference Id shows up verbatim as one of the "/"- or ":"-
# separated tokens inside REMARKS, e.g. a Network Reference Id of
# "00000000J406" appears in a REMARKS value like:
#   "IME-90809428421-845:90809428421/00000000J406:IME REMIT/S43161794"
# and the trailing "/S43161794" is the switch's own ISO/session id for
# that statement line — used to tell whether a given credit into the
# parking account has *already* been reversed back out (a matching DR
# entry elsewhere in the statement carrying the same ISO id).
# ---------------------------------------------------------------------------

_STATEMENT_TOKEN_SPLIT_RE = re.compile(r"[/:]")
_STATEMENT_ISO_ID_RE = re.compile(r"/(S\d+)\s*$")


@dataclass
class BankStatementIndex:
    """In-memory index over a parsed bank statement, built once and reused
    for every row being checked."""

    # normalized token (e.g. a Network Reference Id) -> list of statement
    # rows ({"entry_type", "remarks", "amount", "date"}) whose REMARKS
    # contains that token.
    by_token: dict[str, list[dict]] = field(default_factory=dict)
    # ISO id (e.g. "S43161794") -> set of entry types ("CR"/"DR") seen
    # against that ISO id anywhere in the statement.
    iso_entry_types: dict[str, set[str]] = field(default_factory=dict)
    # Every parsed row, in file order — used where a token can be embedded
    # *inside* a compound "/"- or ":"-delimited token (e.g. NCHL's
    # "SCT-00000000KOQI" contains the account id "00000000KOQI" but isn't
    # equal to it), so a plain by_token lookup would miss it and a
    # substring scan over full REMARKS is needed instead.
    entries: list[dict] = field(default_factory=list)
    total_rows: int = 0


def _read_bank_statement_rows(path: str | Path, source: str = "") -> list[dict]:
    """Read a bank statement export (.csv or .xlsx) into a list of dicts
    with keys entry_type / remarks / amount / date / source. Column names
    are matched case-insensitively (S.N / ENTRY TYPE / REMARKS / AMOUNT /
    DATE / SOURCE).

    `source` is the fallback bank label ("global" / "prabhu") to tag every
    row with when the file itself has no SOURCE column (i.e. a single
    original export, not one of combine_bank_statement_files()'s combined
    outputs) — see BankStatementIndex / statement_entries_for_reference()
    for why this matters (a "failed" row whose Debtor Bank is Prabhu Bank
    must only be checked against Prabhu Bank statement lines, never
    against Global IME Bank ones, and vice versa)."""
    path = Path(path)
    suffix = path.suffix.lower()

    def _normalize_headers(raw_headers):
        return [str(h).strip().upper() if h is not None else "" for h in raw_headers]

    rows: list[dict] = []

    if suffix == ".csv":
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            try:
                header = _normalize_headers(next(reader))
            except StopIteration:
                raise ProcessingError("The bank statement file is empty.")
            try:
                entry_idx = header.index("ENTRY TYPE")
                remarks_idx = header.index("REMARKS")
            except ValueError:
                raise ProcessingError(
                    "The bank statement is missing an 'ENTRY TYPE' or "
                    "'REMARKS' column. Please upload the original bank "
                    "statement export."
                )
            amount_idx = header.index("AMOUNT") if "AMOUNT" in header else None
            date_idx = header.index("DATE") if "DATE" in header else None
            source_idx = header.index("SOURCE") if "SOURCE" in header else None
            for raw in reader:
                if not raw or all(not str(c).strip() for c in raw):
                    continue
                row_source = (
                    raw[source_idx].strip().lower()
                    if source_idx is not None and source_idx < len(raw) and raw[source_idx]
                    else source
                )
                rows.append(
                    {
                        "entry_type": raw[entry_idx].strip() if entry_idx < len(raw) else "",
                        "remarks": raw[remarks_idx].strip() if remarks_idx < len(raw) else "",
                        "amount": raw[amount_idx] if amount_idx is not None and amount_idx < len(raw) else None,
                        "date": raw[date_idx] if date_idx is not None and date_idx < len(raw) else None,
                        "source": row_source,
                    }
                )
    else:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            header = None
            entry_idx = remarks_idx = amount_idx = date_idx = source_idx = None
            for raw in ws.iter_rows(values_only=True):
                if header is None:
                    names = _normalize_headers(raw)
                    if "ENTRY TYPE" in names and "REMARKS" in names:
                        header = names
                        entry_idx = header.index("ENTRY TYPE")
                        remarks_idx = header.index("REMARKS")
                        amount_idx = header.index("AMOUNT") if "AMOUNT" in header else None
                        date_idx = header.index("DATE") if "DATE" in header else None
                        source_idx = header.index("SOURCE") if "SOURCE" in header else None
                    continue
                if raw is None or all(v is None for v in raw):
                    continue
                row_source = (
                    str(raw[source_idx]).strip().lower()
                    if source_idx is not None and source_idx < len(raw) and raw[source_idx]
                    else source
                )
                rows.append(
                    {
                        "entry_type": str(raw[entry_idx]).strip() if entry_idx < len(raw) and raw[entry_idx] is not None else "",
                        "remarks": str(raw[remarks_idx]).strip() if remarks_idx < len(raw) and raw[remarks_idx] is not None else "",
                        "amount": raw[amount_idx] if amount_idx is not None and amount_idx < len(raw) else None,
                        "date": raw[date_idx] if date_idx is not None and date_idx < len(raw) else None,
                        "source": row_source,
                    }
                )
            if header is None:
                raise ProcessingError(
                    "Could not find the bank statement header row (expected "
                    "columns including 'ENTRY TYPE' and 'REMARKS')."
                )
        finally:
            wb.close()

    return rows


def combine_bank_statement_files(paths: list[str | Path], combined_path: str | Path) -> Path:
    """Read one or more bank statement exports (.csv/.xlsx — e.g. up to 4
    separate daily Prabhu Bank statements) and dump every row into a
    single combined .csv file (S.N re-serialized 1, 2, 3, ... across all
    of them), so downstream processing only ever has to open one file.
    Global IME Bank statements normally come as a single file, but this
    works the same way for one file too.

    `paths` may be plain paths (source left blank — the pre-existing,
    untagged behavior) or (source_label, path) tuples, where source_label
    is "global" or "prabhu". Tagging the source lets the failed-sheet
    check later restrict itself to the correct bank's statement lines
    when checking a Prabhu-debtor row vs. a Global-debtor row instead of
    matching against whichever statement happens to be uploaded — see
    statement_entries_for_reference().
    """
    if not paths:
        raise ProcessingError("No bank statement file(s) were provided.")

    combined_path = Path(combined_path)
    combined_path.parent.mkdir(parents=True, exist_ok=True)

    with open(combined_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["S.N", "ENTRY TYPE", "REMARKS", "AMOUNT", "DATE", "SOURCE"])
        sn = 1
        for entry in paths:
            if isinstance(entry, tuple):
                source_label, path = entry
            else:
                source_label, path = "", entry
            for row in _read_bank_statement_rows(path, source=source_label):
                writer.writerow(
                    [
                        sn,
                        row.get("entry_type", ""),
                        row.get("remarks", ""),
                        row.get("amount", ""),
                        row.get("date", ""),
                        row.get("source", "") or source_label,
                    ]
                )
                sn += 1

    return combined_path


def build_bank_statement_index(path: str | Path, default_source: str = "") -> BankStatementIndex:
    rows = _read_bank_statement_rows(path, source=default_source)
    by_token: dict[str, list[dict]] = {}
    iso_entry_types: dict[str, set[str]] = {}

    for row in rows:
        remarks = row["remarks"] or ""
        entry_type = (row["entry_type"] or "").strip().upper()

        tokens = {t.strip().upper() for t in _STATEMENT_TOKEN_SPLIT_RE.split(remarks) if t.strip()}
        for tok in tokens:
            by_token.setdefault(tok, []).append(row)

        iso_match = _STATEMENT_ISO_ID_RE.search(remarks.strip())
        if iso_match:
            iso_id = iso_match.group(1).strip().upper()
            iso_entry_types.setdefault(iso_id, set()).add(entry_type)

    return BankStatementIndex(by_token=by_token, iso_entry_types=iso_entry_types, entries=rows, total_rows=len(rows))


def statement_entries_for_reference(index: BankStatementIndex, ref_id: Any, source: str = "") -> list[dict]:
    """Every bank statement row whose REMARKS contains this (normalized)
    Network Reference Id as a token.

    If `source` is given ("global" or "prabhu"), rows are additionally
    restricted to ones tagged with that same source OR left untagged
    (blank source — e.g. a single, non-combined statement file, where
    there's nothing to disambiguate and every row is implicitly "the
    bank that was uploaded"). This is what stops a Prabhu Bank failed
    row from being matched against a Global IME Bank statement line (or
    vice versa) when both were uploaded together as one combined
    statement."""
    entries = index.by_token.get(normalize_reference_id(ref_id), [])
    if not source:
        return entries
    return [e for e in entries if not e.get("source") or e.get("source") == source]


def is_failed_but_credited(index: BankStatementIndex, ref_id: Any, source: str = "") -> bool:
    """A FAILED-status transaction whose Network Reference Id nonetheless
    shows up in the bank statement means money actually moved (credited
    into the parking account) despite the switch reporting it FAILED —
    this needs a manual reversal just like a normal failed-turned-reversal
    row would.

    `source` restricts the match to statement lines from that bank only
    (see statement_entries_for_reference()) — e.g. a FAILED row whose
    Debtor Bank is Prabhu Bank must only be confirmed "credited" against
    a Prabhu Bank statement line, never a Global IME Bank one that
    happens to share the uploaded (combined) statement."""
    return len(statement_entries_for_reference(index, ref_id, source)) > 0


def is_onus_already_success(index: BankStatementIndex, ref_id: Any, source: str = "") -> bool:
    """On-Us "already successful" check — a looser, name-verified sibling
    of has_duplicate_dr().

    Real-world example that has_duplicate_dr() alone MISSES:

        CR  CEMTFT8897716:7032899459/00000000LH53:City Remit/S54545565
        DR  CEMTFT8897716:BHUNU LAL URAWN:MYS:7032899459/00000000LH53:CEMTFT8897716/S54545574
        DR  TRRR/000012412272/28-07-2026 15:19:30/SCT/S54545565/28-07-2026/S54549126

    The Network Reference Id ("CEMTFT8897716") shows up as a CR (money
    into the parking account) and as exactly ONE DR (the real payout to
    "BHUNU LAL URAWN") — has_duplicate_dr() requires the reference id to
    appear as DR *twice* and so misses this genuinely-successful transfer.
    The third line is a SEPARATE, later system reversal issued (by
    mistake) against the original CR's own trailing ISO id ("S54545565",
    caught by is_already_reversed()) — i.e. support staff reversed a
    transaction that had, in fact, already paid out successfully.

    So: this reference id shows up as a DR entry at all, AND that same DR
    entry's REMARKS also carries what looks like a genuine beneficiary
    name (has letters, more than one word — not just a numeric/settlement
    token) — same "reference id *with a name*" verification NCHL's
    is_already_debited_nchl() uses, adapted for On-Us's single shared
    reference id instead of NCHL's separate CR/DR anchors.

    `source` restricts the match to statement lines from that bank only
    — see statement_entries_for_reference()."""
    entries = statement_entries_for_reference(index, ref_id, source)
    if not any((e["entry_type"] or "").strip().upper() == "CR" for e in entries):
        return False
    for entry in entries:
        if (entry["entry_type"] or "").strip().upper() != "DR":
            continue
        remarks = entry["remarks"] or ""
        for tok in _STATEMENT_TOKEN_SPLIT_RE.split(remarks):
            tok = tok.strip()
            if not tok:
                continue
            # A genuine beneficiary name: multiple words, letters only
            # (spaces allowed) — filters out settlement/account ids, ISO
            # ids, and short alpha codes like "MYS"/"SCT".
            words = tok.split()
            if len(words) >= 2 and all(w.replace(".", "").isalpha() for w in words):
                return True
    return False


def has_duplicate_dr(index: BankStatementIndex, ref_id: Any, source: str = "") -> bool:
    """True if this Network Reference Id shows up as a DR (debit) entry
    two or more times in the statement.

    Only meaningful for On-Us rows (Debtor Bank and Creditor Bank both our
    own Global IME Bank): an on-us transfer settles entirely within the
    bank and never lands in the parking/settlement account the way an
    off-us transaction's CR leg does, so it won't be caught by
    is_failed_but_credited() the normal way. If a FAILED-status on-us row's
    Network Reference Id nonetheless shows up as DR *twice*, that's both
    legs of the transfer actually completing end-to-end — i.e. it already
    succeeded, in full, and must NOT be treated as needing a manual
    reversal (see apply_bank_statement_to_reversal_file()'s "already
    succeeded (On-Us)" handling, which is checked before, and instead of,
    the ordinary failed-but-credited path).

    `source` restricts the match to statement lines from that bank only
    — see statement_entries_for_reference()."""
    entries = statement_entries_for_reference(index, ref_id, source)
    dr_count = sum(1 for e in entries if (e["entry_type"] or "").strip().upper() == "DR")
    return dr_count >= 2


def is_already_reversed(index: BankStatementIndex, ref_id: Any, source: str = "") -> bool:
    """A manual-reversal row's Network Reference Id should show up as a CR
    entry in the bank statement (the original credit into the parking
    account). That CR entry's REMARKS ends with the switch's own ISO id
    (e.g. "/S43161794"). If that same ISO id also shows up anywhere inside
    a DR entry's REMARKS elsewhere in the statement (not necessarily at
    the end of it — a reversal's own REMARKS typically has the *original*
    ISO id somewhere in the middle, followed by the reversal transaction's
    own trailing id), the money has already gone back out — i.e. support
    (or the system) already reversed it, and this row on the manual
    reversal file would be a double reversal.

    `source` restricts the initial CR lookup to statement lines from that
    bank only — see statement_entries_for_reference(). The secondary
    ISO-id chase (to find the matching DR) is left unrestricted since a
    reversal's own DR entry lives in the same bank's statement as its CR
    counterpart by construction."""
    entries = statement_entries_for_reference(index, ref_id, source)
    for entry in entries:
        if (entry["entry_type"] or "").strip().upper() != "CR":
            continue
        iso_match = _STATEMENT_ISO_ID_RE.search((entry["remarks"] or "").strip())
        if not iso_match:
            continue
        iso_id = iso_match.group(1).strip().upper()
        # The ISO id is itself a "/"- or ":"-delimited token, so it was
        # already indexed by build_bank_statement_index() the same way a
        # Network Reference Id is — look it up directly rather than only
        # checking other rows' *trailing* ISO id.
        for candidate in index.by_token.get(iso_id, []):
            if (candidate["entry_type"] or "").strip().upper() == "DR":
                return True
    return False


_STATEMENT_ACCOUNT_TOKEN_RE = re.compile(r"0{5,}[A-Z0-9]{2,}")


def is_already_debited_nchl(index: BankStatementIndex, ref_id: Any, source: str = "") -> bool:
    """NCHL-specific "already handled" check.

    For most networks, a reversal's DR entry carries the *original* CR
    entry's own trailing ISO id somewhere in its REMARKS, so
    is_already_reversed() (matching on that ISO id) catches a double
    reversal. NCHL settlement statements are different: the CR and DR legs
    for the very same underlying payment don't share the Network Reference
    Id *or* the ISO id at all — e.g.:

        CR  SCT/00000000KOQI/24883884901595327/11001000/DURLAV MAN SHRESTHA/S51287331
        DR  CIPS/SCT-00000000KOQI/DURLAV MAN SHRESTHA/15373/691590444/IPS/S51287339

    The two legs only reliably share two things: the beneficiary name
    (the token immediately before the CR's own trailing ISO id — "DURLAV
    MAN SHRESTHA" above) and the masked settlement account id embedded in
    both REMARKS ("00000000KOQI" above — present standalone in the CR leg,
    and embedded inside the compound token "SCT-00000000KOQI" in the DR
    leg, hence a substring scan rather than a token-equality lookup).

    Matching on name alone is too loose — the same beneficiary can
    legitimately receive several unrelated payments, so a name-only match
    risks red-flagging (and thus silently not reversing) an unrelated,
    still-outstanding transaction. So this requires *both* anchors — name
    **and** the settlement account id — to agree on the same DR entry
    before treating it as already debited/settled.

    One more wrinkle: that matched DR entry is itself a statement line
    with its own trailing ISO id. If *that* ISO id also shows up as a CR
    entry elsewhere in the statement, NCHL has effectively reversed its
    own reversal — the money that left as this DR came right back in, so
    the original credit is back to sitting unreversed in the parking
    account. In that case this DR is not a terminal "already debited"
    state after all, and the row must still go through the normal manual
    reversal (i.e. this function must NOT report it as already-debited).
    """
    entries = statement_entries_for_reference(index, ref_id, source)
    for entry in entries:
        if (entry["entry_type"] or "").strip().upper() != "CR":
            continue
        remarks = entry["remarks"] or ""
        tokens = [t.strip() for t in _STATEMENT_TOKEN_SPLIT_RE.split(remarks) if t.strip()]
        if len(tokens) < 2:
            continue
        # The token right before the CR's own trailing ISO id — normally
        # the beneficiary name, e.g. "DURLAV MAN SHRESTHA".
        anchor_name = tokens[-2].strip().upper()
        if not anchor_name or not any(ch.isalpha() for ch in anchor_name):
            # Skip purely numeric anchors (account/session numbers) — too
            # generic and prone to matching an unrelated DR entry.
            continue

        # The masked settlement account id (e.g. "00000000KOQI") — the
        # second, more specific anchor. Without it we can't safely
        # disambiguate same-name transactions, so skip this entry rather
        # than risk a false-positive match on name alone.
        account_match = _STATEMENT_ACCOUNT_TOKEN_RE.search(remarks.upper())
        if not account_match:
            continue
        anchor_account = account_match.group(0)

        for candidate in index.entries:
            if candidate is entry:
                continue
            if (candidate["entry_type"] or "").strip().upper() != "DR":
                continue
            candidate_remarks = (candidate["remarks"] or "").upper()
            if anchor_name not in candidate_remarks or anchor_account not in candidate_remarks:
                continue

            # This DR looks like the "already debited" leg — but check
            # whether NCHL itself reversed *this* DR back out (its own
            # trailing ISO id also shows up as a CR entry elsewhere). If
            # so, the money came back in and this isn't a terminal
            # already-debited state — don't flag it, let it go through
            # the normal manual reversal instead.
            candidate_iso_match = _STATEMENT_ISO_ID_RE.search((candidate["remarks"] or "").strip())
            if candidate_iso_match:
                candidate_iso = candidate_iso_match.group(1).strip().upper()
                # Mirrors is_already_reversed()'s own chase: a CR entry that
                # undoes this DR will carry the DR's trailing ISO id
                # somewhere inside its REMARKS (not necessarily trailing),
                # so look it up via by_token (every occurrence), not
                # iso_entry_types (which only tracks *trailing* ids).
                reversed_back = any(
                    (row["entry_type"] or "").strip().upper() == "CR"
                    for row in index.by_token.get(candidate_iso, [])
                )
                if reversed_back:
                    continue

            return True
    return False


@dataclass
class BankStatementCheckStats:
    statement_rows: int = 0
    failed_credited_count: int = 0
    failed_credited_amount: float = 0.0
    failed_credited_charge: float = 0.0
    already_reversed_count: int = 0
    already_reversed_amount: float = 0.0
    already_reversed_charge: float = 0.0
    already_reversed_by_sheet: dict[str, int] = field(default_factory=dict)
    # Subset of already_reversed_* above that was caught specifically by
    # the NCHL direct-reference-id (CR-alongside-DR) check rather than the
    # general ISO-id chase — tracked separately so it's visible that this
    # bug fix is actually catching rows.
    already_reversed_nchl_count: int = 0
    already_reversed_nchl_amount: float = 0.0
    # Subset of already_reversed_* caught by the On-Us "already successful"
    # (duplicate DR) check — see has_duplicate_dr(). This catches an On-Us
    # (Global-to-Global or Prabhu-to-Prabhu) row that ended up in a manual
    # reversal sheet (coop/imeremit/cityremit/prabhu) by mistake even
    # though it already completed successfully end-to-end, verified the
    # same way the "failed" sheet's On-Us check works.
    already_reversed_onus_success_count: int = 0
    already_reversed_onus_success_amount: float = 0.0
    # Rows on the ONUS_SYSTEM_REVERSAL_SHEET_NAME sheet (On-Us/NCHL system
    # reversals — see process_ibft_file()) that turn out to have already
    # completed successfully BEFORE the system reversal was issued (see
    # is_onus_already_success()) — i.e. support staff reversed a
    # transaction that had, in fact, already paid out. Red-flagged on that
    # sheet so these get reviewed/escalated rather than treated as routine.
    onus_system_reversal_flagged_count: int = 0
    onus_system_reversal_flagged_amount: float = 0.0
    # On-Us/Off-Us split of the failed-but-credited rows above, plus their
    # reason labels — used to subtract these back out of the "pure failed"
    # on-us/off-us breakdown on the log (a credited row isn't a pure
    # failure any more, so it shouldn't count toward that breakdown).
    failed_credited_onus_count: int = 0
    failed_credited_offus_count: int = 0
    failed_credited_onus_amount: float = 0.0
    failed_credited_offus_amount: float = 0.0
    failed_credited_reason_onus: dict[str, int] = field(default_factory=dict)
    failed_credited_reason_offus: dict[str, int] = field(default_factory=dict)
    # On-Us rows found genuinely already-successful via a duplicate DR (see
    # has_duplicate_dr()) — pulled out of "failed" entirely, and kept
    # separate from failed_credited_* above since these do NOT need a
    # manual reversal (unlike a failed-but-credited row).
    onus_already_success_count: int = 0
    onus_already_success_amount: float = 0.0
    onus_already_success_charge: float = 0.0
    onus_already_success_reason: dict[str, int] = field(default_factory=dict)

    def as_dict(self) -> dict[str, Any]:
        return {
            "failed_credited_count": self.failed_credited_count,
            "failed_credited_amount": self.failed_credited_amount,
            "failed_credited_charge": self.failed_credited_charge,
            "already_reversed_count": self.already_reversed_count,
            "already_reversed_amount": self.already_reversed_amount,
            "already_reversed_charge": self.already_reversed_charge,
            "already_reversed_nchl_count": self.already_reversed_nchl_count,
            "already_reversed_nchl_amount": self.already_reversed_nchl_amount,
            "already_reversed_onus_success_count": self.already_reversed_onus_success_count,
            "already_reversed_onus_success_amount": self.already_reversed_onus_success_amount,
            "onus_system_reversal_flagged_count": self.onus_system_reversal_flagged_count,
            "onus_system_reversal_flagged_amount": self.onus_system_reversal_flagged_amount,
            "onus_already_success_count": self.onus_already_success_count,
            "onus_already_success_amount": self.onus_already_success_amount,
            "onus_already_success_charge": self.onus_already_success_charge,
        }


_RED_FLAG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GREEN_FLAG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _flag_row_red(ws, row_number):
    for cell in ws[row_number]:
        cell.fill = _RED_FLAG_FILL
        existing = cell.font
        cell.font = Font(bold=bool(existing and existing.bold), color="9C0006")


def _flag_row_green(ws, row_number):
    for cell in ws[row_number]:
        cell.fill = _GREEN_FLAG_FILL
        existing = cell.font
        cell.font = Font(bold=bool(existing and existing.bold), color="006100")


def _expected_statement_source(debtor_bank: Any) -> str:
    """Which bank statement a row's Network Reference Id should be looked
    up against, based on its Debtor Bank — "prabhu" for Prabhu Bank,
    "global" for Global IME Bank, "" (no restriction — check every
    statement line regardless of source) for anything else/unrecognized,
    so processing still works for a bank that hasn't been tagged."""
    if is_prabhu_bank(debtor_bank):
        return "prabhu"
    if is_global_bank(debtor_bank):
        return "global"
    return ""


def apply_bank_statement_to_reversal_file(
    reversal_file_path: str | Path,
    statement_path: str | Path,
    default_source: str = "",
) -> BankStatementCheckStats:
    """Re-open an already-generated need_to_reversal_*.xlsx and cross-check
    it against a bank statement export:

      - "failed" sheet: any row whose Network Reference Id shows up in the
        statement gets red-flagged in place, and is also copied into a new
        "Failed but Credited" sheet laid out like the reversal sheets
        (same derived Debit/Credit Account Number + Narration columns)
        since it needs to be manually reversed. On-Us rows (Debtor Bank
        and Creditor Bank both our own Global IME Bank) are checked
        first for a duplicate DR (see has_duplicate_dr()) — if found, the
        transfer actually completed successfully end-to-end and does NOT
        need a reversal, so it's green-flagged instead and copied into a
        separate "Already Success (OnUs)" sheet (plain columns, no
        reversal-row transform) rather than "Failed but Credited".
      - "coop" / "imeremit" / "cityremit" sheets: any row whose original
        credit has already been reversed out (per `is_already_reversed`)
        gets red-flagged in place, so whoever works the file can see it
        was already handled and skip it (rather than double-reversing).

    The workbook is saved back to `reversal_file_path` in place. Rows are
    never deleted or reordered — only highlighted — so nothing about the
    file's existing structure changes for anyone already relying on it.

    `default_source` tags every row of `statement_path` with this bank
    ("global" or "prabhu") when the file itself has no per-row SOURCE
    column — i.e. a single, non-combined statement upload (only ever
    one bank's data, so every line implicitly belongs to it). This is
    what lets the Prabhu-vs-Global restriction below work correctly even
    for the common case of uploading just one statement at a time; a
    combined multi-file upload is already tagged row-by-row by
    combine_bank_statement_files(), so this default is only a fallback.
    """
    index = build_bank_statement_index(statement_path, default_source=default_source)

    try:
        wb = openpyxl.load_workbook(reversal_file_path)
    except Exception as exc:
        raise ProcessingError(f"Could not open the generated reversal file: {exc}") from exc

    stats = BankStatementCheckStats(statement_rows=index.total_rows)

    # --- "failed" sheet -----------------------------------------------
    failed_but_credited_rows: list[tuple] = []
    onus_already_success_rows: list[tuple] = []
    if "failed" in wb.sheetnames:
        ws = wb["failed"]
        header = [c.value for c in ws[1]]
        col = {name: i for i, name in enumerate(header) if name}
        if "Network Reference Id" in col:
            ref_idx = col["Network Reference Id"]
            for row_number in range(2, ws.max_row + 1):
                values = [ws.cell(row=row_number, column=c + 1).value for c in range(len(header))]
                if all(v is None for v in values):
                    continue
                ref_id = values[ref_idx]
                if not ref_id:
                    continue

                onus = (
                    "Debtor Bank" in col
                    and "Creditor Bank" in col
                    and is_on_us(values[col["Debtor Bank"]], values[col["Creditor Bank"]])
                )
                # Restrict matching to the statement for this row's own
                # bank (e.g. a Prabhu-debtor FAILED row's CR must be
                # checked against the Prabhu Bank statement specifically,
                # not a Global IME Bank one that happens to be part of
                # the same combined upload) — see
                # _expected_statement_source().
                expected_source = _expected_statement_source(
                    values[col["Debtor Bank"]] if "Debtor Bank" in col else None
                )

                # On-Us + a duplicate DR = both legs of the transfer already
                # completed end-to-end — genuinely successful, not stuck
                # waiting on a reversal. Checked (and handled) separately
                # from, and before, the ordinary failed-but-credited path
                # below, since these rows must NOT be treated as needing a
                # manual reversal.
                if onus and (
                    has_duplicate_dr(index, ref_id, expected_source)
                    or is_onus_already_success(index, ref_id, expected_source)
                ):
                    _flag_row_green(ws, row_number)
                    stats.onus_already_success_count += 1
                    if "Transaction Amount" in col:
                        stats.onus_already_success_amount += to_float(values[col["Transaction Amount"]])
                    if "Charge Amount" in col and "Payment Processor" in col:
                        stats.onus_already_success_charge += to_float(resolve_charge_amount(tuple(values), col))
                    if "Source Message" in col:
                        reason = normalize_failure_reason(values[col["Source Message"]])
                        stats.onus_already_success_reason[reason] = (
                            stats.onus_already_success_reason.get(reason, 0) + 1
                        )
                    onus_already_success_rows.append(tuple(values))
                    continue

                if not is_failed_but_credited(index, ref_id, expected_source):
                    continue
                _flag_row_red(ws, row_number)
                stats.failed_credited_count += 1
                if "Transaction Amount" in col:
                    stats.failed_credited_amount += to_float(values[col["Transaction Amount"]])
                if "Charge Amount" in col and "Payment Processor" in col:
                    stats.failed_credited_charge += to_float(resolve_charge_amount(tuple(values), col))
                if "Debtor Bank" in col and "Creditor Bank" in col:
                    debtor_bank = values[col["Debtor Bank"]]
                    creditor_bank = values[col["Creditor Bank"]]
                    reason = normalize_failure_reason(
                        values[col["Source Message"]] if "Source Message" in col else None
                    )
                    if is_on_us(debtor_bank, creditor_bank):
                        stats.failed_credited_onus_count += 1
                        if "Transaction Amount" in col:
                            stats.failed_credited_onus_amount += to_float(values[col["Transaction Amount"]])
                        stats.failed_credited_reason_onus[reason] = (
                            stats.failed_credited_reason_onus.get(reason, 0) + 1
                        )
                    else:
                        stats.failed_credited_offus_count += 1
                        if "Transaction Amount" in col:
                            stats.failed_credited_offus_amount += to_float(values[col["Transaction Amount"]])
                        stats.failed_credited_reason_offus[reason] = (
                            stats.failed_credited_reason_offus.get(reason, 0) + 1
                        )
                failed_but_credited_rows.append(tuple(values))

    raw_col = {name: i for i, name in enumerate(REQUIRED_COLUMNS)}

    if onus_already_success_rows:
        # Plain original-columns layout (like "failed" itself) — these rows
        # don't need the reversal-row transform since no reversal is due.
        if "Already Success (OnUs)" in wb.sheetnames:
            del wb["Already Success (OnUs)"]
        _write_raw_sheet(wb, "Already Success (OnUs)", onus_already_success_rows, raw_col)

    if failed_but_credited_rows:
        # Reuses the same "REQUIRED_COLUMNS-ordered raw row" -> reversal-row
        # transform used for the normal reversal sheets, since the
        # "failed" sheet is written with that same 23-column layout.
        if "Failed but Credited" in wb.sheetnames:
            del wb["Failed but Credited"]
        _write_reversal_sheet_flat(wb, "Failed but Credited", failed_but_credited_rows, raw_col)

    # --- coop / imeremit / cityremit sheets -----------------------------
    for sheet_name in _REVERSAL_SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = None
        col: dict[str, int] = {}
        for row_number in range(1, ws.max_row + 1):
            values = [ws.cell(row=row_number, column=c + 1).value for c in range(ws.max_column)]
            if all(v is None for v in values):
                header = None  # blank separator row between coop member groups
                continue
            if values[0] == "S NO":
                header = values
                col = {name: i for i, name in enumerate(header) if name}
                continue
            if header is None or "Network Reference Id" not in col:
                continue
            ref_id = values[col["Network Reference Id"]]
            if not ref_id:
                continue

            expected_source = _expected_statement_source(
                values[col["Debtor Bank"]] if "Debtor Bank" in col else None
            )

            already = is_already_reversed(index, ref_id, expected_source)
            is_nchl_hit = False
            is_onus_success_hit = False

            if (
                not already
                and "Debtor Bank" in col
                and "Creditor Bank" in col
                and is_on_us(values[col["Debtor Bank"]], values[col["Creditor Bank"]])
            ):
                # On-Us (Global-to-Global or Prabhu-to-Prabhu) rows can end
                # up on a manual-reversal sheet by mistake even though both
                # legs of the transfer already completed successfully
                # end-to-end. Verify success the same way the "failed"
                # sheet's On-Us check does (a duplicate DR entry — see
                # has_duplicate_dr()) and, if confirmed, red-flag it here so
                # nobody reverses an already-successful transaction a
                # second time.
                if has_duplicate_dr(index, ref_id, expected_source) or is_onus_already_success(index, ref_id, expected_source):
                    already = True
                    is_onus_success_hit = True

            if not already and "Payment Processor" in col and is_zero_charge_network(values[col["Payment Processor"]]):
                # NCHL bug fix: the CR/DR legs for the same reference id can
                # carry two different trailing ISO ids, so the general
                # ISO-id chase above misses it — check the reference id
                # directly for a CR-alongside-DR pair instead.
                if is_already_debited_nchl(index, ref_id, expected_source):
                    already = True
                    is_nchl_hit = True

            if not already:
                continue
            _flag_row_red(ws, row_number)
            stats.already_reversed_count += 1
            stats.already_reversed_by_sheet[sheet_name] = stats.already_reversed_by_sheet.get(sheet_name, 0) + 1
            if "Transaction Amount" in col:
                stats.already_reversed_amount += to_float(values[col["Transaction Amount"]])
            if "Charge Amount" in col:
                stats.already_reversed_charge += to_float(values[col["Charge Amount"]])
            if is_nchl_hit:
                stats.already_reversed_nchl_count += 1
                if "Transaction Amount" in col:
                    stats.already_reversed_nchl_amount += to_float(values[col["Transaction Amount"]])
            if is_onus_success_hit:
                stats.already_reversed_onus_success_count += 1
                if "Transaction Amount" in col:
                    stats.already_reversed_onus_success_amount += to_float(values[col["Transaction Amount"]])

    # --- On-Us/NCHL system reversal sheet -------------------------------
    # These rows were already system-reversed by the switch (or, as seen
    # in practice, by support staff manually triggering a reversal) — but
    # some of them turn out to have already completed successfully BEFORE
    # that reversal happened (see is_onus_already_success()'s docstring for
    # a real example). Red-flag those here so they get reviewed/escalated
    # instead of looking like a routine system reversal.
    if ONUS_SYSTEM_REVERSAL_SHEET_NAME in wb.sheetnames:
        ws = wb[ONUS_SYSTEM_REVERSAL_SHEET_NAME]
        header = [c.value for c in ws[1]]
        col = {name: i for i, name in enumerate(header) if name}
        if "Network Reference Id" in col:
            ref_idx = col["Network Reference Id"]
            for row_number in range(2, ws.max_row + 1):
                values = [ws.cell(row=row_number, column=c + 1).value for c in range(len(header))]
                if all(v is None for v in values):
                    continue
                ref_id = values[ref_idx]
                if not ref_id:
                    continue

                expected_source = _expected_statement_source(
                    values[col["Debtor Bank"]] if "Debtor Bank" in col else None
                )

                already_success = has_duplicate_dr(index, ref_id, expected_source) or is_onus_already_success(
                    index, ref_id, expected_source
                )
                if not already_success and "Payment Processor" in col and is_zero_charge_network(values[col["Payment Processor"]]):
                    already_success = is_already_debited_nchl(index, ref_id, expected_source)

                if not already_success:
                    continue
                _flag_row_red(ws, row_number)
                stats.onus_system_reversal_flagged_count += 1
                if "Transaction Amount" in col:
                    stats.onus_system_reversal_flagged_amount += to_float(values[col["Transaction Amount"]])

    Path(reversal_file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(reversal_file_path)
    wb.close()

    return stats


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_THIN_SIDE = Side(style="thin", color="000000")
_ALL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _autofit(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)


def _style_header_row(ws, row_number, total_columns, highlight_trailing=0):
    """Bold + bordered header row; optionally highlight the last N columns
    (yellow fill) to call out generated/derived fields."""
    highlight_from = total_columns - highlight_trailing + 1
    for cell in ws[row_number]:
        cell.font = _HEADER_FONT
        cell.border = _ALL_BORDER
        if cell.column >= highlight_from:
            cell.fill = _YELLOW_FILL


def _style_data_row(ws, row_number):
    for cell in ws[row_number]:
        cell.border = _ALL_BORDER


# Columns that must always be written/kept as TEXT, never as a number.
# Long numeric-looking ids (17+ digits) silently lose trailing precision if
# Excel is allowed to store them as a number — this was the root cause of a
# past merchant-id mismatch between the source file and the generated
# reversal file, so every id-like column is force-formatted as text.
_ID_LIKE_COLUMN_NAMES = {
    "Member Transaction Id",
    "Network Reference Id",
    "Session Id",
    "Debit Account Number",
    "Credit Account Number",
}


def _force_text_cell(row_number, ws, col_idx):
    cell = ws.cell(row=row_number, column=col_idx)
    if cell.value is not None:
        cell.value = str(cell.value)
    cell.number_format = "@"


def _apply_id_text_format(ws, row_number, headers):
    for idx, name in enumerate(headers, start=1):
        if name in _ID_LIKE_COLUMN_NAMES:
            _force_text_cell(row_number, ws, idx)


def _write_raw_sheet(wb, sheet_name, rows, col):
    """Write rows using the original 23-column layout, with S NO
    re-serialized (1, 2, 3, ...) rather than the original file's S NO."""
    ws = wb.create_sheet(sheet_name)
    ws.append(REQUIRED_COLUMNS)
    _style_header_row(ws, ws.max_row, len(REQUIRED_COLUMNS))

    for i, row in enumerate(rows, start=1):
        values = [row[col[name]] for name in REQUIRED_COLUMNS]
        values[0] = i  # re-serialize S NO
        ws.append(values)
        _style_data_row(ws, ws.max_row)
        _apply_id_text_format(ws, ws.max_row, REQUIRED_COLUMNS)

    _autofit(ws)
    return ws


def _reversal_row_values(row, col, s_no):
    def g(name):
        return row[col[name]]

    narration = build_narration(g("Member Transaction Id"), g("Session Id"))
    charge_amount = resolve_charge_amount(row, col)
    debit_account, _recognized = resolve_debit_account(g("Debtor Bank"))
    return [
        s_no,
        g("Transaction Date"),
        g("Member Name"),
        g("Aggregator"),
        # Member Transaction Id / Network Reference Id are written as text
        # further down (_force_text_columns) so Excel never reinterprets a
        # long numeric-looking id as a number (which silently rounds/
        # truncates it past ~15 significant digits and was the cause of a
        # past merchant-id mismatch).
        g("Member Transaction Id"),
        g("Network Reference Id"),
        g("Session Id"),
        g("Payment Processor"),
        g("Transaction Amount"),
        charge_amount,
        g("Debtor Bank"),
        g("Creditor Bank"),
        debit_account,
        g("Debit Account Number"),  # money goes back to where it came from
        narration,
    ]


def _write_reversal_sheet_flat(wb, sheet_name, rows, col):
    """imeremit / cityremit: one continuous block, original relative order
    preserved (source file is already most-recent-first)."""
    ws = wb.create_sheet(sheet_name)
    ws.append(REVERSAL_SHEET_HEADERS)
    _style_header_row(ws, ws.max_row, len(REVERSAL_SHEET_HEADERS), HIGHLIGHTED_TRAILING_COLUMNS)

    for i, row in enumerate(rows, start=1):
        ws.append(_reversal_row_values(row, col, i))
        _style_data_row(ws, ws.max_row)
        _apply_id_text_format(ws, ws.max_row, REVERSAL_SHEET_HEADERS)

    _autofit(ws)
    return ws


def _write_reversal_sheet_grouped_by_member(wb, sheet_name, rows, col, stats):
    """coop: grouped by Member Name (case-insensitive alphabetical order),
    each group gets its own header row and S NO restarting at 1, separated
    by a single blank row."""
    ws = wb.create_sheet(sheet_name)

    groups: dict[str, list] = {}
    for row in rows:
        member = str(row[col["Member Name"]] or "").strip()
        groups.setdefault(member, []).append(row)

    stats.coop_member_count = len(groups)

    ordered_members = sorted(groups.keys(), key=lambda m: m.lower())

    first_group = True
    for member in ordered_members:
        if not first_group:
            ws.append([None] * len(REVERSAL_SHEET_HEADERS))
        first_group = False

        ws.append(REVERSAL_SHEET_HEADERS)
        _style_header_row(ws, ws.max_row, len(REVERSAL_SHEET_HEADERS), HIGHLIGHTED_TRAILING_COLUMNS)

        for i, row in enumerate(groups[member], start=1):
            ws.append(_reversal_row_values(row, col, i))
            _style_data_row(ws, ws.max_row)
            _apply_id_text_format(ws, ws.max_row, REVERSAL_SHEET_HEADERS)

    _autofit(ws)
    return ws


# ---------------------------------------------------------------------------
# Output filename helper
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _extract_date_str(source_filename: str) -> str:
    match = _DATE_RE.search(source_filename)
    return match.group(1) if match else date.today().isoformat()


def build_output_filename(source_filename: str) -> str:
    """need_to_reversal_<date>.xlsx"""
    return f"need_to_reversal_{_extract_date_str(source_filename)}.xlsx"


def build_uploaded_filename(source_filename: str) -> str:
    """ibft_txn_data_<date>.xlsx — every uploaded source file is renamed to
    this standardized name (date taken from the original filename if it
    contains one, otherwise today's date) so the audit trail is consistent
    regardless of what the file was originally called."""
    return f"ibft_txn_data_{_extract_date_str(source_filename)}.xlsx"
