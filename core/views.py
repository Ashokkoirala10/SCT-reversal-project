import io
import time
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files import File
from django.core.paginator import Paginator
from django.db.models import Count, Max, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import BankAccountForm, BankStatementUploadForm, UploadForm
from .models import BankAccount, MemberAggregatorStat, ProcessingLog, SeenNetworkReferenceId
from .services import (
    ProcessingError,
    apply_bank_statement_to_reversal_file,
    build_output_filename,
    build_uploaded_filename,
    combine_bank_statement_files,
    extract_reversal_network_reference_ids,
    process_ibft_file,
)

PAGE_SIZE = 10


def is_admin(user):
    return user.is_authenticated and user.is_staff


def can_toggle_passed(user, log):
    return user.is_staff or user.username == log.uploaded_by


class BrandedLoginView(auth_views.LoginView):
    template_name = "core/login.html"
    redirect_authenticated_user = True


@login_required
def upload_view(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = form.cleaned_data["ibft_file"]

            # Standardize the stored/displayed name regardless of what the
            # file was actually called when it was uploaded.
            standardized_name = build_uploaded_filename(uploaded.name)

            log = ProcessingLog.objects.create(
                uploaded_file=uploaded,
                uploaded_filename=standardized_name,
                uploaded_by=request.user.username,
            )
            input_path = Path(log.uploaded_file.path)
            output_filename = build_output_filename(uploaded.name)
            tmp_output_path = Path(settings.MEDIA_ROOT) / "outputs" / f"tmp_{log.id}_{output_filename}"
            tmp_output_path.parent.mkdir(parents=True, exist_ok=True)

            # --- Double-reversal prevention -----------------------------
            # Look at the most recently generated file that has already
            # been reviewed and marked PASSED (not just any generated
            # file), and pull out every Network Reference Id it already
            # reversed, so this run can skip them.
            previous_log = (
                ProcessingLog.objects.filter(
                    status=ProcessingLog.STATUS_SUCCESS,
                    passed=True,
                    generated_file__isnull=False,
                )
                .exclude(id=log.id)
                .order_by("-created_at")
                .first()
            )
            previously_reversed_ids = set()
            if previous_log and previous_log.generated_file:
                try:
                    previously_reversed_ids = extract_reversal_network_reference_ids(
                        previous_log.generated_file.path
                    )
                except Exception:
                    previously_reversed_ids = set()

            # --- Overlapping-time-window dedup ---------------------------
            # Every Network Reference Id already processed out of an
            # earlier upload that was reviewed and marked PASSED — not
            # *any* earlier upload — so a row from a period already
            # covered by yesterday's approved download doesn't get counted
            # (or reversed) a second time today. Files that were generated
            # but never marked passed (e.g. a bad run that got
            # regenerated) must not block today's run from seeing those
            # same rows again.
            seen_reference_ids = set(
                SeenNetworkReferenceId.objects.filter(source_log__passed=True).values_list(
                    "ref_id", flat=True
                )
            )

            t0 = time.perf_counter()
            try:
                stats = process_ibft_file(
                    input_path, tmp_output_path, previously_reversed_ids, seen_reference_ids
                )
            except ProcessingError as exc:
                log.status = ProcessingLog.STATUS_FAILED
                log.error_message = str(exc)
                log.save()
                return render(
                    request,
                    "core/upload.html",
                    {"form": form, "error": str(exc), **_panel_context(request)},
                )
            except Exception as exc:  # noqa: BLE001 - surface unexpected errors to the audit log too
                log.status = ProcessingLog.STATUS_FAILED
                log.error_message = f"Unexpected error: {exc}"
                log.save()
                return render(
                    request,
                    "core/upload.html",
                    {
                        "form": form,
                        "error": f"Unexpected error while processing the file: {exc}",
                        **_panel_context(request),
                    },
                )
            elapsed_ms = int((time.perf_counter() - t0) * 1000)

            with open(tmp_output_path, "rb") as f:
                log.generated_file.save(output_filename, File(f), save=False)
            tmp_output_path.unlink(missing_ok=True)

            log.generated_filename = output_filename
            log.status = ProcessingLog.STATUS_SUCCESS
            log.processing_duration_ms = elapsed_ms
            kept_reference_ids = stats.kept_reference_ids
            member_aggregator_breakdown = stats.member_aggregator_breakdown
            for key, value in stats.as_dict().items():
                setattr(log, key, value)
            log.save()

            # Persist the per (Member Name, Aggregator) breakdown captured
            # during processing — powers the dashboard's Member &
            # Aggregator report (success/failed/reversal count + amount).
            if member_aggregator_breakdown:
                magg_objs = []
                for key, bucket in member_aggregator_breakdown.items():
                    member_name, _, aggregator = key.partition("\x1f")
                    magg_objs.append(
                        MemberAggregatorStat(
                            log=log,
                            member_name=member_name,
                            aggregator=aggregator,
                            success_count=bucket["success_count"],
                            success_amount=bucket["success_amount"],
                            failed_count=bucket["failed_count"],
                            failed_amount=bucket["failed_amount"],
                            reversal_count=bucket["reversal_count"],
                            reversal_amount=bucket["reversal_amount"],
                        )
                    )
                MemberAggregatorStat.objects.bulk_create(magg_objs, batch_size=500)

            # Persist newly-seen Network Reference Ids for the next
            # upload's overlapping-window dedup check.
            new_refs = [
                SeenNetworkReferenceId(ref_id=r, source_log=log)
                for r in kept_reference_ids
                if r not in seen_reference_ids
            ]
            if new_refs:
                SeenNetworkReferenceId.objects.bulk_create(
                    new_refs, ignore_conflicts=True, batch_size=1000
                )

            return redirect(reverse("core:result", args=[log.id]))
    else:
        form = UploadForm()

    return render(request, "core/upload.html", {"form": form, **_panel_context(request)})


@login_required
def bank_statement_upload_view(request):
    if request.method == "POST":
        form = BankStatementUploadForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            log = form.cleaned_data["target_log"]
            global_files = form.cleaned_data["global_statement_files"]
            prabhu_files = form.cleaned_data["prabhu_statement_files"]

            if not can_toggle_passed(request.user, log):
                messages.error(request, "You can only check your own generated files against a bank statement.")
                return redirect("core:bank_statement_upload")

            tmp_dir = Path(settings.MEDIA_ROOT) / "bank_statements"
            tmp_dir.mkdir(parents=True, exist_ok=True)

            if global_files and prabhu_files:
                bank_type = ProcessingLog.BANK_TYPE_BOTH
            elif prabhu_files:
                bank_type = ProcessingLog.BANK_TYPE_PRABHU
            else:
                bank_type = ProcessingLog.BANK_TYPE_GLOBAL

            saved_paths = []
            saved_names = []
            for prefix, statement in [("global", f) for f in global_files] + [("prabhu", f) for f in prabhu_files]:
                tmp_path = tmp_dir / f"{log.id}_{prefix}_{statement.name}"
                with open(tmp_path, "wb") as fh:
                    for chunk in statement.chunks():
                        fh.write(chunk)
                saved_paths.append((prefix, tmp_path))
                saved_names.append(statement.name)

            # One file: used as-is. Multiple files (Global + up to 4 Prabhu
            # exports, any combination) — dump them all into one combined
            # file first (tagging each row with which bank it came from —
            # see combine_bank_statement_files()) so the rest of the
            # pipeline only ever deals with a single statement, checking
            # DR/CR the same way either way, while still being able to
            # restrict a Prabhu-debtor row to Prabhu statement lines only
            # (and a Global-debtor row to Global statement lines only).
            if len(saved_paths) > 1:
                combined_path = tmp_dir / f"{log.id}_{bank_type}_combined.csv"
                try:
                    statement_path = combine_bank_statement_files(saved_paths, combined_path)
                except ProcessingError as exc:
                    messages.error(request, str(exc))
                    return redirect("core:bank_statement_upload")
            else:
                statement_path = saved_paths[0][1]

            try:
                default_source = {
                    ProcessingLog.BANK_TYPE_GLOBAL: "global",
                    ProcessingLog.BANK_TYPE_PRABHU: "prabhu",
                }.get(bank_type, "")
                bstats = apply_bank_statement_to_reversal_file(log.generated_file.path, statement_path, default_source)
            except ProcessingError as exc:
                messages.error(request, str(exc))
                return redirect("core:bank_statement_upload")
            except Exception as exc:  # noqa: BLE001
                messages.error(request, f"Unexpected error while checking the bank statement: {exc}")
                return redirect("core:bank_statement_upload")

            log.bank_statement_checked = True
            log.bank_statement_filename = ", ".join(saved_names)
            log.bank_statement_type = bank_type
            log.bank_statement_checked_by = request.user.username
            log.bank_statement_checked_at = timezone.now()
            for key, value in bstats.as_dict().items():
                setattr(log, key, value)

            # A credited row isn't a "pure" failure any more, so pull it
            # back out of the on-us/off-us failed breakdown (count and
            # reason) it was originally bucketed into.
            log.failed_onus_count = max(0, log.failed_onus_count - bstats.failed_credited_onus_count)
            log.failed_offus_count = max(0, log.failed_offus_count - bstats.failed_credited_offus_count)
            log.failed_onus_amount = max(0.0, log.failed_onus_amount - bstats.failed_credited_onus_amount)
            log.failed_offus_amount = max(0.0, log.failed_offus_amount - bstats.failed_credited_offus_amount)
            for reason, count in bstats.failed_credited_reason_onus.items():
                current = dict(log.failed_reason_breakdown_onus or {})
                current[reason] = max(0, current.get(reason, 0) - count)
                log.failed_reason_breakdown_onus = current
            for reason, count in bstats.failed_credited_reason_offus.items():
                current = dict(log.failed_reason_breakdown_offus or {})
                current[reason] = max(0, current.get(reason, 0) - count)
                log.failed_reason_breakdown_offus = current

            # An On-Us row with a duplicate DR isn't "failed" at all any
            # more — it's genuinely successful (both legs of the transfer
            # completed). Pull it out of every failed bucket (total, on-us,
            # and reason breakdown) and add it to success instead.
            if bstats.onus_already_success_count:
                log.failed_total = max(0, log.failed_total - bstats.onus_already_success_count)
                log.failed_amount = max(0.0, log.failed_amount - bstats.onus_already_success_amount)
                log.failed_onus_count = max(0, log.failed_onus_count - bstats.onus_already_success_count)
                log.failed_onus_amount = max(0.0, log.failed_onus_amount - bstats.onus_already_success_amount)
                log.success_count = log.success_count + bstats.onus_already_success_count
                log.success_amount = log.success_amount + bstats.onus_already_success_amount
                for reason, count in bstats.onus_already_success_reason.items():
                    current = dict(log.failed_reason_breakdown_onus or {})
                    current[reason] = max(0, current.get(reason, 0) - count)
                    log.failed_reason_breakdown_onus = current
                    current_all = dict(log.failed_reason_breakdown or {})
                    current_all[reason] = max(0, current_all.get(reason, 0) - count)
                    log.failed_reason_breakdown = current_all

            log.save()

            bank_label = dict(ProcessingLog.BANK_TYPE_CHOICES).get(bank_type, bank_type)
            nchl_note = (
                f" (incl. {bstats.already_reversed_nchl_count} caught by the NCHL ref-id check)"
                if bstats.already_reversed_nchl_count else ""
            )
            khalti_note = (
                f" (incl. {bstats.already_reversed_khalti_count} caught by the Khalti settlement check)"
                if bstats.already_reversed_khalti_count else ""
            )
            onus_manual_note = (
                f" (incl. {bstats.already_reversed_onus_success_count} On-Us row(s) on the manual "
                f"reversal sheets found already successful and red-flagged)"
                if bstats.already_reversed_onus_success_count else ""
            )
            onus_success_note = (
                f" {bstats.onus_already_success_count} On-Us row(s) found already successful (duplicate DR) "
                f"and moved to success."
                if bstats.onus_already_success_count else ""
            )
            onus_sys_rev_note = (
                f" {bstats.onus_system_reversal_flagged_count} On-Us/NCHL system reversal row(s) found to have "
                f"already succeeded before the reversal was issued and red-flagged for review."
                if bstats.onus_system_reversal_flagged_count else ""
            )
            messages.success(
                request,
                f"Checked against {bank_label} statement ({', '.join(saved_names)}): "
                f"{bstats.failed_credited_count} failed-but-credited row(s) and "
                f"{bstats.already_reversed_count} already-reversed row(s) red-flagged{nchl_note}{khalti_note}{onus_manual_note}."
                f"{onus_success_note}{onus_sys_rev_note}",
            )
            return redirect(reverse("core:result", args=[log.id]))
    else:
        form = BankStatementUploadForm(user=request.user)

    context = {"form": form}
    if is_admin(request.user):
        # "Extra" page's second tab — add-bank-account feature for Admin
        # (is_staff) users. Full CRUD on bank accounts (edit/delete) stays
        # superuser-only via Django admin (see core/admin.py); this is
        # just a quick "add a new one" form plus a read-only reference
        # list of what's currently configured.
        context["bank_form"] = BankAccountForm()
        context["bank_accounts"] = BankAccount.objects.order_by("bank_name")
    return render(request, "core/bank_statement_upload.html", context)


@login_required
@user_passes_test(is_admin, login_url="core:upload")
@require_POST
def add_bank_account_view(request):
    """Add a new Debtor-Bank -> Debit-Account mapping (core.models.BankAccount)
    from the "Extra" page's Add bank account tab. Restricted to Admin
    (is_staff) users. Editing or deleting an existing row still requires
    Django admin (superuser) — see core/admin.py's BankAccountAdmin, which
    is left exactly as it was."""
    form = BankAccountForm(request.POST)
    if form.is_valid():
        account = form.save()
        messages.success(
            request,
            f"Bank account '{account.bank_name}' (keyword '{account.keyword}') added.",
        )
    else:
        for field, errors in form.errors.items():
            label = form.fields[field].label if field in form.fields else field
            for err in errors:
                messages.error(request, f"{label}: {err}")
    return redirect("core:bank_statement_upload")


def _panel_context(request):
    """Context for the two activity panels on the main page:
    - `shared_page`: central "passed" reports, visible to everyone
    - `mine_page`: the current user's own activity (every status), with a
      toggle to mark a generated file as passed so it graduates into the
      shared list and into the analytics dashboard.
    """
    shared_logs = ProcessingLog.objects.filter(passed=True)
    shared_paginator = Paginator(shared_logs, PAGE_SIZE)
    shared_page = shared_paginator.get_page(request.GET.get("page"))

    mine_logs = ProcessingLog.objects.filter(uploaded_by=request.user.username)
    mine_paginator = Paginator(mine_logs, PAGE_SIZE)
    mine_page = mine_paginator.get_page(request.GET.get("mypage"))

    return {"shared_page": shared_page, "mine_page": mine_page}


@login_required
def result_view(request, log_id):
    log = get_object_or_404(ProcessingLog, id=log_id)
    unique_new_txns = max(0, log.total_rows - log.duplicate_source_skipped)
    return render(
        request,
        "core/result.html",
        {"log": log, "can_toggle": can_toggle_passed(request.user, log), "unique_new_txns": unique_new_txns},
    )


@login_required
@require_POST
def toggle_passed_view(request, log_id):
    log = get_object_or_404(ProcessingLog, id=log_id)

    if not can_toggle_passed(request.user, log):
        messages.error(request, "You can only mark your own reports as passed.")
        return redirect(request.META.get("HTTP_REFERER") or reverse("core:upload"))

    if log.status != ProcessingLog.STATUS_SUCCESS:
        messages.error(request, "Only a successfully generated file can be marked as passed.")
        return redirect(request.META.get("HTTP_REFERER") or reverse("core:upload"))

    log.passed = not log.passed
    if log.passed:
        log.passed_by = request.user.username
        log.passed_at = timezone.now()
    else:
        log.passed_by = ""
        log.passed_at = None
    log.save(update_fields=["passed", "passed_by", "passed_at"])

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or reverse("core:upload")
    return redirect(next_url)


@login_required
def download_file_view(request, log_id, kind):
    log = get_object_or_404(ProcessingLog, id=log_id)

    if kind == "uploaded":
        field, filename = log.uploaded_file, log.uploaded_filename
    elif kind == "generated":
        field, filename = log.generated_file, log.generated_filename
    else:
        raise Http404("Unknown file kind")

    if not field:
        raise Http404("File not available")

    # Serve with the clean, standardized filename regardless of whatever
    # suffix Django's storage may have appended to the file on disk (e.g.
    # need_to_reversal_2026-07-10_5WKRS2i.xlsx when a same-named file
    # already existed) — the download should always be need_to_reversal_2026-07-10.xlsx.
    return FileResponse(field.open("rb"), as_attachment=True, filename=filename or field.name)


def _reclassify_log(log):
    """Same red-flag reclassification used on the dashboard, applied per
    row: failed-but-credited rows move out of "failed" and into manual
    reversal (money moved, still needs reversing); already-reversed rows
    move out of manual reversal and into system reversal (support/system
    already moved the money back)."""
    log.failed_excl_credited = max(0, log.failed_total - log.failed_credited_count)
    log.manual_reversal_clean = (
        max(0, log.reversal_manual_kept - log.already_reversed_count) + log.failed_credited_count
    )
    log.system_reversal_clean = log.reversal_system_count + log.already_reversed_count
    return log


@login_required
@user_passes_test(is_admin, login_url="core:upload")
def audit_log_view(request):
    logs = ProcessingLog.objects.all()
    paginator = Paginator(logs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page"))
    for log in page_obj:
        _reclassify_log(log)
    return render(request, "core/audit_log.html", {"page_obj": page_obj})


@login_required
@user_passes_test(is_admin, login_url="core:upload")
def export_audit_log_view(request):
    logs = ProcessingLog.objects.all().order_by("-created_at")
    wb, ws = _new_sheet("Audit log")
    headers = [
        "#", "When", "Uploaded by", "Source file", "Generated file", "Status", "Passed",
        "Manual reversal (incl. credited)", "System reversal (incl. already-reversed)",
        "Failed (excl. credited)", "Timeout", "Prabhu rerouted", "Unrecognized bank rows",
        "Duplicates skipped",
    ]
    row = _write_table_header(ws, 1, headers)
    for log in logs:
        _reclassify_log(log)
        ws.append([
            log.id,
            timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M") if log.created_at else "",
            log.uploaded_by or "—",
            log.uploaded_filename or "—",
            log.generated_filename or "—",
            log.status,
            "Passed" if log.passed else ("Pending review" if log.status == ProcessingLog.STATUS_SUCCESS else "—"),
            log.manual_reversal_clean,
            log.system_reversal_clean,
            log.failed_excl_credited,
            log.timeout_count,
            log.prabhu_rerouted,
            log.unrecognized_debtor_bank_rows,
            log.duplicate_skipped,
        ])
    _style_data_rows(ws, row, len(headers))
    _autosize(ws)
    return _finalize_xlsx(wb, "audit_log.xlsx")


_MONTH_NAMES = [
    (1, "January"), (2, "February"), (3, "March"), (4, "April"),
    (5, "May"), (6, "June"), (7, "July"), (8, "August"),
    (9, "September"), (10, "October"), (11, "November"), (12, "December"),
]

_REASON_ORDER = [
    "Timeout",
    "Insufficient fund",
    "Card issuer timeout",
    "Response timeout",
    "Transaction amount exceeded",
    "Other",
]


def _parse_date_range(request, prefix: str):
    """Reads `{prefix}_from` / `{prefix}_to` (YYYY-MM-DD) GET params —
    used by the Member-wise and Aggregator-wise report tabs so each can be
    narrowed to a specific date range independently of the page-level
    year/month filter. Returns (from_date, to_date), either possibly None."""
    from datetime import datetime as _datetime

    def _parse(raw):
        raw = (raw or "").strip()
        if not raw:
            return None
        try:
            return _datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            return None

    return _parse(request.GET.get(f"{prefix}_from")), _parse(request.GET.get(f"{prefix}_to"))


def _apply_year_month_filter(request):
    """Shared by the dashboard and every CSV export: filters passed
    ProcessingLogs down to the selected year/month (either/both/neither),
    and returns everything needed to render the filter controls too."""
    successful_logs = ProcessingLog.objects.filter(status=ProcessingLog.STATUS_SUCCESS, passed=True)

    available_years = sorted(
        {
            timezone.localtime(dt).year
            for dt in successful_logs.values_list("created_at", flat=True)
            if dt
        },
        reverse=True,
    )
    selected_year_raw = (request.GET.get("year") or "").strip()
    selected_year = int(selected_year_raw) if selected_year_raw.isdigit() else None
    if selected_year:
        successful_logs = successful_logs.filter(created_at__year=selected_year)

    selected_month_raw = (request.GET.get("month") or "").strip()
    selected_month = int(selected_month_raw) if selected_month_raw.isdigit() and 1 <= int(selected_month_raw) <= 12 else None
    selected_month_name = dict(_MONTH_NAMES).get(selected_month) if selected_month else None
    if selected_month:
        successful_logs = successful_logs.filter(created_at__month=selected_month)

    # Day-of-month filter (dd) — combines with month/year above (either/
    # both/neither) so the dashboard can be narrowed all the way down to a
    # single calendar date (dd/mm/yyyy), not just a whole month or year.
    selected_day_raw = (request.GET.get("day") or "").strip()
    selected_day = int(selected_day_raw) if selected_day_raw.isdigit() and 1 <= int(selected_day_raw) <= 31 else None
    if selected_day:
        successful_logs = successful_logs.filter(created_at__day=selected_day)

    return {
        "logs": successful_logs,
        "available_years": available_years,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_month_name": selected_month_name,
        "selected_day": selected_day,
        "month_names": _MONTH_NAMES,
        "day_numbers": list(range(1, 32)),
    }


def _compute_totals(successful_logs):
    """Shared headline totals + red-flag reclassification math, used by
    the dashboard and the summary CSV export."""
    totals = successful_logs.aggregate(
        files=Count("id"),
        total_rows=Sum("total_rows"),
        success_count=Sum("success_count"),
        success_amount=Sum("success_amount"),
        failed_total=Sum("failed_total"),
        failed_kept=Sum("failed_kept"),
        failed_amount=Sum("failed_amount"),
        failed_charge=Sum("failed_charge"),
        failed_kept_amount=Sum("failed_kept_amount"),
        failed_kept_charge=Sum("failed_kept_charge"),
        reversal_manual_kept=Sum("reversal_manual_kept"),
        reversal_manual_amount=Sum("reversal_manual_amount"),
        reversal_manual_charge=Sum("reversal_manual_charge"),
        reversal_system_count=Sum("reversal_system_count"),
        reversal_system_amount=Sum("reversal_system_amount"),
        reversal_system_charge=Sum("reversal_system_charge"),
        coop_count=Sum("coop_count"),
        imeremit_count=Sum("imeremit_count"),
        cityremit_count=Sum("cityremit_count"),
        timeout_count=Sum("timeout_count"),
        timeout_amount=Sum("timeout_amount"),
        prabhu_rerouted=Sum("prabhu_rerouted"),
        duplicate_skipped=Sum("duplicate_skipped"),
        duplicate_source_skipped=Sum("duplicate_source_skipped"),
        unrecognized_debtor_bank_rows=Sum("unrecognized_debtor_bank_rows"),
        failed_credited_count=Sum("failed_credited_count"),
        failed_credited_amount=Sum("failed_credited_amount"),
        failed_credited_charge=Sum("failed_credited_charge"),
        already_reversed_count=Sum("already_reversed_count"),
        already_reversed_amount=Sum("already_reversed_amount"),
        already_reversed_charge=Sum("already_reversed_charge"),
        failed_onus_count=Sum("failed_onus_count"),
        failed_onus_amount=Sum("failed_onus_amount"),
        failed_offus_count=Sum("failed_offus_count"),
        failed_offus_amount=Sum("failed_offus_amount"),
    )
    for key, value in totals.items():
        if value is None:
            totals[key] = 0

    totals["failed_clean_count"] = max(0, totals["failed_total"] - totals["failed_credited_count"])
    totals["failed_clean_amount"] = round(totals["failed_amount"] - totals["failed_credited_amount"], 2)
    totals["manual_reversal_clean_count"] = max(
        0, totals["reversal_manual_kept"] - totals["already_reversed_count"]
    )
    totals["manual_reversal_clean_amount"] = round(
        totals["reversal_manual_amount"] - totals["already_reversed_amount"], 2
    )
    totals["system_reversal_clean_count"] = totals["reversal_system_count"] + totals["already_reversed_count"]
    totals["system_reversal_clean_amount"] = round(
        totals["reversal_system_amount"] + totals["already_reversed_amount"], 2
    )
    totals["total_reversal_count"] = (
        totals["manual_reversal_clean_count"]
        + totals["system_reversal_clean_count"]
        + totals["failed_credited_count"]
    )
    return totals


def _ordered_reasons(bucket: dict, total: int) -> list:
    rows = []
    for reason in _REASON_ORDER:
        count = bucket.get(reason, 0)
        rows.append({"reason": reason, "count": count, "pct": round(100 * count / total) if total else 0})
    for reason, count in bucket.items():
        if reason not in _REASON_ORDER:
            rows.append({"reason": reason, "count": count, "pct": round(100 * count / total) if total else 0})
    return rows


def _compute_onus_offus(successful_logs, totals):
    """Shared on-us/off-us failed breakdown, used by the dashboard and
    the on-us/off-us CSV exports."""
    onus_reason_totals: dict = {}
    offus_reason_totals: dict = {}
    for log in successful_logs.only("id", "failed_reason_breakdown_onus", "failed_reason_breakdown_offus"):
        for reason, count in (log.failed_reason_breakdown_onus or {}).items():
            onus_reason_totals[reason] = onus_reason_totals.get(reason, 0) + count
        for reason, count in (log.failed_reason_breakdown_offus or {}).items():
            offus_reason_totals[reason] = offus_reason_totals.get(reason, 0) + count

    return {
        "onus": {
            "count": totals["failed_onus_count"],
            "amount": round(totals["failed_onus_amount"], 2),
            "reasons": _ordered_reasons(onus_reason_totals, totals["failed_onus_count"]),
        },
        "offus": {
            "count": totals["failed_offus_count"],
            "amount": round(totals["failed_offus_amount"], 2),
            "reasons": _ordered_reasons(offus_reason_totals, totals["failed_offus_count"]),
        },
    }


def _build_daily_stats(successful_logs):
    """Day-by-day breakdown, red-flag-corrected, with full success/failed/
    manual-reversal/system-reversal/failed-credited/already-reversed/timeout
    counts, amounts AND charges, plus top failure reasons, processing time
    and "data through" — exactly what the dashboard's day cards and their
    "click for full reconciliation" modal show (each item's "detail" key is
    literally what's rendered there, via {{ d.detail|json_script:... }}).

    Shared by dashboard_view (on-screen, paginated) and
    export_day_breakdown_view (the Excel download) so the two can never
    drift apart — the download always matches the website, not a
    simplified one-line-per-day summary.
    """
    daily_qs = (
        successful_logs.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            files=Count("id"),
            total_rows=Sum("total_rows"),
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_total=Sum("failed_total"),
            failed_kept=Sum("failed_kept"),
            failed_insufficient_funds=Sum("failed_insufficient_funds"),
            failed_amount=Sum("failed_amount"),
            failed_charge=Sum("failed_charge"),
            reversal_manual_kept=Sum("reversal_manual_kept"),
            reversal_manual_amount=Sum("reversal_manual_amount"),
            reversal_manual_charge=Sum("reversal_manual_charge"),
            reversal_system_count=Sum("reversal_system_count"),
            reversal_system_amount=Sum("reversal_system_amount"),
            reversal_system_charge=Sum("reversal_system_charge"),
            timeout_count=Sum("timeout_count"),
            timeout_amount=Sum("timeout_amount"),
            duplicate_skipped=Sum("duplicate_skipped"),
            duplicate_source_skipped=Sum("duplicate_source_skipped"),
            prabhu_rerouted=Sum("prabhu_rerouted"),
            unrecognized_debtor_bank_rows=Sum("unrecognized_debtor_bank_rows"),
            failed_credited_count=Sum("failed_credited_count"),
            failed_credited_amount=Sum("failed_credited_amount"),
            already_reversed_count=Sum("already_reversed_count"),
            already_reversed_amount=Sum("already_reversed_amount"),
            processing_duration_ms=Sum("processing_duration_ms"),
            data_through_at=Max("data_through_at"),
        )
        .order_by("-day")
    )

    # Pull each day's per-log failure-reason breakdowns and merge them.
    reason_by_day: dict = {}
    for log in successful_logs.exclude(failed_reason_breakdown={}):
        day = log.created_at.date()
        bucket = reason_by_day.setdefault(day, {})
        for reason, count in (log.failed_reason_breakdown or {}).items():
            bucket[reason] = bucket.get(reason, 0) + count

    daily_stats = []
    chart_points = []  # oldest -> newest, for the line graph
    for row in daily_qs:
        failed_credited = row["failed_credited_count"] or 0
        already_reversed = row["already_reversed_count"] or 0
        failed = max(0, (row["failed_total"] or 0) - failed_credited)
        reversal = max(0, (row["reversal_manual_kept"] or 0) - already_reversed)
        system_reversal = (row["reversal_system_count"] or 0) + already_reversed
        total_reversal = reversal + system_reversal + failed_credited
        timeout = row["timeout_count"] or 0
        generated_total = (row["success_count"] or 0) + failed + reversal + timeout
        failed_total_all = row["failed_total"] or 0
        day = row["day"]

        reasons = reason_by_day.get(day, {})
        top_reasons = sorted(reasons.items(), key=lambda kv: kv[1], reverse=True)[:6]
        top_reasons_pct = [
            {
                "reason": reason,
                "count": count,
                "pct": round(100 * count / failed_total_all) if failed_total_all else 0,
            }
            for reason, count in top_reasons
        ]

        total_rows = row["total_rows"] or 0
        dup_source_skipped = row["duplicate_source_skipped"] or 0
        detail = {
            "day": day.strftime("%A, %d %b %Y"),
            "day_iso": day.isoformat(),
            "files": row["files"],
            "total_rows": total_rows,
            "unique_new_txns": max(0, total_rows - dup_source_skipped),
            "success": {
                "count": row["success_count"] or 0,
                "amount": round(row["success_amount"] or 0, 2),
            },
            "failed": {
                "count": failed,
                "amount": round((row["failed_amount"] or 0) - (row["failed_credited_amount"] or 0), 2),
                "charge": round(row["failed_charge"] or 0, 2),
            },
            "manual_reversal": {
                "count": reversal,
                "amount": round((row["reversal_manual_amount"] or 0) - (row["already_reversed_amount"] or 0), 2),
                "charge": round(row["reversal_manual_charge"] or 0, 2),
            },
            "system_reversal": {
                "count": system_reversal,
                "amount": round((row["reversal_system_amount"] or 0) + (row["already_reversed_amount"] or 0), 2),
                "charge": round(row["reversal_system_charge"] or 0, 2),
            },
            "failed_credited": {
                "count": failed_credited,
                "amount": round(row["failed_credited_amount"] or 0, 2),
            },
            "already_reversed": {
                "count": already_reversed,
                "amount": round(row["already_reversed_amount"] or 0, 2),
            },
            "timeout": {
                "count": timeout,
                "amount": round(row["timeout_amount"] or 0, 2),
            },
            "duplicate_skipped": row["duplicate_skipped"] or 0,
            "duplicate_source_skipped": row["duplicate_source_skipped"] or 0,
            "prabhu_rerouted": row["prabhu_rerouted"] or 0,
            "unrecognized_debtor_bank_rows": row["unrecognized_debtor_bank_rows"] or 0,
            "top_reasons": top_reasons_pct,
            "processing_ms": row["processing_duration_ms"] or 0,
            "data_through_at": timezone.localtime(row["data_through_at"]).strftime("%Y-%m-%d %I:%M %p")
            if row["data_through_at"]
            else None,
        }

        daily_stats.append(
            {
                "day": day,
                "day_id": f"day-detail-{day.isoformat()}",
                "files": row["files"],
                "total_rows": total_rows,
                "unique_new_txns": detail["unique_new_txns"],
                "success_count": row["success_count"] or 0,
                "failed_kept": failed,
                "failed_total_all": failed_total_all,
                "failed_insufficient_funds": row["failed_insufficient_funds"] or 0,
                "reversal_manual_kept": reversal,
                "reversal_system_count": system_reversal,
                "total_reversal": total_reversal,
                "timeout_count": timeout,
                "duplicate_skipped": row["duplicate_skipped"] or 0,
                "duplicate_source_skipped": row["duplicate_source_skipped"] or 0,
                "processing_ms": row["processing_duration_ms"] or 0,
                "data_through_at": detail["data_through_at"],
                "success_pct": round(100 * (row["success_count"] or 0) / generated_total) if generated_total else 0,
                "failed_pct": round(100 * failed / generated_total) if generated_total else 0,
                "reversal_pct": round(100 * total_reversal / generated_total) if generated_total else 0,
                "timeout_pct": round(100 * timeout / generated_total) if generated_total else 0,
                "top_reasons": top_reasons_pct,
                "detail": detail,
            }
        )
        chart_points.append(
            {
                "day": day.strftime("%d %b"),
                "success": row["success_count"] or 0,
                "failed": failed,
                "reversal": total_reversal,
            }
        )

    chart_points.reverse()  # oldest first for the line graph
    return daily_stats, chart_points


@login_required
def dashboard_view(request):
    # Only PASSED files count toward analytics — a generated file that
    # hasn't been reviewed/marked passed yet doesn't skew the numbers.
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    available_years = year_month["available_years"]
    selected_year = year_month["selected_year"]
    selected_month = year_month["selected_month"]
    selected_month_name = year_month["selected_month_name"]
    selected_day = year_month["selected_day"]
    month_names = year_month["month_names"]
    day_numbers = year_month["day_numbers"]

    totals = _compute_totals(successful_logs)

    # --- Failed on-us / off-us breakdown, with reasons ---------------------
    # On-Us = Debtor Bank is our own bank (Global IME Bank); Off-Us = any
    # other bank. Reason labels come pre-bucketed (case-insensitively) into
    # the fixed set: Timeout, Insufficient fund, Card issuer timeout,
    # Response timeout, Transaction amount exceeded, Other.
    onus_offus = _compute_onus_offus(successful_logs, totals)

    # --- Monthly report ------------------------------------------------
    monthly_qs = (
        successful_logs.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(
            files=Count("id"),
            total_rows=Sum("total_rows"),
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_total=Sum("failed_total"),
            failed_amount=Sum("failed_amount"),
            reversal_manual_kept=Sum("reversal_manual_kept"),
            reversal_manual_amount=Sum("reversal_manual_amount"),
            reversal_system_count=Sum("reversal_system_count"),
            reversal_system_amount=Sum("reversal_system_amount"),
            timeout_count=Sum("timeout_count"),
            failed_credited_count=Sum("failed_credited_count"),
            failed_credited_amount=Sum("failed_credited_amount"),
            already_reversed_count=Sum("already_reversed_count"),
            already_reversed_amount=Sum("already_reversed_amount"),
        )
        .order_by("-month")
    )
    monthly_stats = []
    for row in monthly_qs:
        failed_credited = row["failed_credited_count"] or 0
        already_reversed = row["already_reversed_count"] or 0
        failed = max(0, (row["failed_total"] or 0) - failed_credited)
        reversal = max(0, (row["reversal_manual_kept"] or 0) - already_reversed)
        system_reversal = (row["reversal_system_count"] or 0) + already_reversed
        total_reversal = reversal + system_reversal + failed_credited
        month_dt = row["month"]
        monthly_stats.append(
            {
                "label": timezone.localtime(month_dt).strftime("%B %Y") if month_dt else "—",
                "files": row["files"],
                "total_rows": row["total_rows"] or 0,
                "success_count": row["success_count"] or 0,
                "success_amount": round(row["success_amount"] or 0, 2),
                "failed_count": failed,
                "failed_amount": round((row["failed_amount"] or 0) - (row["failed_credited_amount"] or 0), 2),
                "manual_reversal_count": reversal,
                "system_reversal_count": system_reversal,
                "total_reversal": total_reversal,
                "timeout_count": row["timeout_count"] or 0,
            }
        )

    daily_stats, chart_points = _build_daily_stats(successful_logs)

    daily_paginator = Paginator(daily_stats, PAGE_SIZE)
    daily_page = daily_paginator.get_page(request.GET.get("page"))

    # --- Member-wise report ----------------------------------------------
    # Success / failed / manual-reversal count + amount, rolled up per
    # Member Name (across every aggregator) over every passed file matching
    # the selected year/month filter, further narrowed by its own optional
    # From/To date range. Optional free-text filter on member name. Kept as
    # its own table (not mixed with the aggregator-wise report below) since
    # the two answer different questions.
    member_query = (request.GET.get("member_q") or "").strip()
    member_from, member_to = _parse_date_range(request, "member")
    member_qs = MemberAggregatorStat.objects.filter(log__in=successful_logs)
    if member_from:
        member_qs = member_qs.filter(log__created_at__date__gte=member_from)
    if member_to:
        member_qs = member_qs.filter(log__created_at__date__lte=member_to)
    member_qs = (
        member_qs.values("member_name")
        .annotate(
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_count=Sum("failed_count"),
            failed_amount=Sum("failed_amount"),
            reversal_count=Sum("reversal_count"),
            reversal_amount=Sum("reversal_amount"),
        )
        .order_by("member_name")
    )
    if member_query:
        member_qs = member_qs.filter(member_name__icontains=member_query)

    member_stats = [
        {
            "member_name": row["member_name"] or "\u2014",
            "success_count": row["success_count"] or 0,
            "success_amount": round(row["success_amount"] or 0, 2),
            "failed_count": row["failed_count"] or 0,
            "failed_amount": round(row["failed_amount"] or 0, 2),
            "reversal_count": row["reversal_count"] or 0,
            "reversal_amount": round(row["reversal_amount"] or 0, 2),
        }
        for row in member_qs
    ]
    member_paginator = Paginator(member_stats, PAGE_SIZE)
    member_page = member_paginator.get_page(request.GET.get("member_page"))

    # --- Aggregator-wise report -------------------------------------------
    # Same breakdown, rolled up per Aggregator (across every member)
    # instead, with its own independent From/To date range. Optional
    # free-text filter on aggregator name.
    aggregator_query = (request.GET.get("aggregator_q") or "").strip()
    aggregator_from, aggregator_to = _parse_date_range(request, "aggregator")
    aggregator_qs = MemberAggregatorStat.objects.filter(log__in=successful_logs)
    if aggregator_from:
        aggregator_qs = aggregator_qs.filter(log__created_at__date__gte=aggregator_from)
    if aggregator_to:
        aggregator_qs = aggregator_qs.filter(log__created_at__date__lte=aggregator_to)
    aggregator_qs = (
        aggregator_qs.values("aggregator")
        .annotate(
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_count=Sum("failed_count"),
            failed_amount=Sum("failed_amount"),
            reversal_count=Sum("reversal_count"),
            reversal_amount=Sum("reversal_amount"),
        )
        .order_by("aggregator")
    )
    if aggregator_query:
        aggregator_qs = aggregator_qs.filter(aggregator__icontains=aggregator_query)

    aggregator_stats = [
        {
            "aggregator": row["aggregator"] or "\u2014",
            "success_count": row["success_count"] or 0,
            "success_amount": round(row["success_amount"] or 0, 2),
            "failed_count": row["failed_count"] or 0,
            "failed_amount": round(row["failed_amount"] or 0, 2),
            "reversal_count": row["reversal_count"] or 0,
            "reversal_amount": round(row["reversal_amount"] or 0, 2),
        }
        for row in aggregator_qs
    ]
    aggregator_paginator = Paginator(aggregator_stats, PAGE_SIZE)
    aggregator_page = aggregator_paginator.get_page(request.GET.get("aggregator_page"))

    return render(
        request,
        "core/dashboard.html",
        {
            "totals": totals,
            "daily_page": daily_page,
            "chart_points": chart_points[-30:],
            "monthly_stats": monthly_stats,
            "onus_offus": onus_offus,
            "available_years": available_years,
            "selected_year": selected_year,
            "month_names": month_names,
            "selected_month": selected_month,
            "selected_month_name": selected_month_name,
            "selected_day": selected_day,
            "day_numbers": day_numbers,
            "member_page": member_page,
            "member_query": member_query,
            "member_from": member_from.isoformat() if member_from else "",
            "member_to": member_to.isoformat() if member_to else "",
            "aggregator_page": aggregator_page,
            "aggregator_query": aggregator_query,
            "aggregator_from": aggregator_from.isoformat() if aggregator_from else "",
            "aggregator_to": aggregator_to.isoformat() if aggregator_to else "",
        },
    )


_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(name="Calibri", bold=True, size=13, color="1D4ED8")
_LABEL_FONT = Font(name="Calibri", bold=True, size=10.5)
_TITLE_FONT = Font(name="Calibri", bold=True, size=15, color="1D4ED8")
_DEFAULT_FONT = Font(name="Calibri", size=10.5)


def _new_sheet(title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws


def _write_table_header(ws, row: int, headers: list) -> int:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"
    return row + 1


def _style_data_rows(ws, header_row: int, num_cols: int) -> None:
    for row in ws.iter_rows(min_row=header_row + 1, max_col=num_cols):
        for cell in row:
            cell.font = _DEFAULT_FONT


def _autosize(ws) -> None:
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 42)


def _write_title(ws, text: str, period: str) -> int:
    ws.cell(row=1, column=1, value=text).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period}").font = _DEFAULT_FONT
    return 4


def _write_section(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    return row + 1


def _write_kv(ws, row: int, label: str, value) -> int:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    v = ws.cell(row=row, column=2, value=value)
    v.font = _DEFAULT_FONT
    return row + 1


def _finalize_xlsx(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _period_suffix(selected_year, selected_month_name, selected_day=None) -> str:
    if selected_month_name and selected_year:
        base = f"{selected_month_name}_{selected_year}"
    elif selected_year:
        base = str(selected_year)
    elif selected_month_name:
        base = selected_month_name
    else:
        base = "all_time"
    if selected_day:
        return f"{base}_{selected_day:02d}" if base != "all_time" else f"day_{selected_day:02d}"
    return base


def _period_label(selected_year, selected_month_name, selected_day=None) -> str:
    if selected_month_name and selected_year:
        base = f"{selected_month_name} {selected_year}"
    elif selected_year:
        base = str(selected_year)
    elif selected_month_name:
        base = f"{selected_month_name} (all years)"
    else:
        base = "All time"
    if selected_day:
        return f"{base}, day {selected_day}" if base not in ("All time",) else f"Day {selected_day} of every month"
    return base



@login_required
def export_failed_onoffus_view(request):
    """One workbook, two sheets — On-Us and Off-Us — so a single download
    covers both failed breakdowns instead of two separate files."""
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    totals = _compute_totals(successful_logs)
    onus_offus = _compute_onus_offus(successful_logs, totals)
    period = _period_label(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))
    suffix = _period_suffix(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))

    wb, ws_onus = _new_sheet("On-Us")
    ws_offus = wb.create_sheet("Off-Us")

    for ws, side_key, label in ((ws_onus, "onus", "Failed — On-Us"), (ws_offus, "offus", "Failed — Off-Us")):
        bucket = onus_offus[side_key]
        row = _write_title(ws, label, period)
        row = _write_kv(ws, row, "Total failed rows", bucket["count"])
        row = _write_kv(ws, row, "Total failed amount (Rs.)", round(bucket["amount"], 2))
        row += 1
        header_row = _write_table_header(ws, row, ["Reason", "Count", "% of bucket"])
        row = header_row - 1
        for r in bucket["reasons"]:
            row += 1
            ws.cell(row=row, column=1, value=r["reason"])
            ws.cell(row=row, column=2, value=r["count"])
            ws.cell(row=row, column=3, value=f"{r['pct']}%")
        _style_data_rows(ws, header_row, 3)
        _autosize(ws)

    return _finalize_xlsx(wb, f"failed_onus_offus_{suffix}.xlsx")


@login_required
def export_dashboard_summary_view(request):
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    totals = _compute_totals(successful_logs)
    onus_offus = _compute_onus_offus(successful_logs, totals)
    period = _period_label(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))
    suffix = _period_suffix(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))

    monthly_qs = (
        successful_logs.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(
            files=Count("id"),
            total_rows=Sum("total_rows"),
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_total=Sum("failed_total"),
            failed_amount=Sum("failed_amount"),
            reversal_manual_kept=Sum("reversal_manual_kept"),
            reversal_system_count=Sum("reversal_system_count"),
            timeout_count=Sum("timeout_count"),
            failed_credited_count=Sum("failed_credited_count"),
            already_reversed_count=Sum("already_reversed_count"),
        )
        .order_by("-month")
    )

    wb, ws = _new_sheet("Summary")
    row = _write_title(ws, "Analytics dashboard summary", period)

    row = _write_section(ws, row, "Volume")
    row = _write_kv(ws, row, "Total files processed", totals["files"])
    row = _write_kv(ws, row, "Total successful transactions", totals["success_count"])
    row = _write_kv(ws, row, "Successful amount (Rs.)", round(totals["success_amount"], 2))
    row = _write_kv(ws, row, "Timeout rows", totals["timeout_count"])
    row += 1

    row = _write_section(ws, row, "Failed")
    row = _write_kv(ws, row, "Total failed (excl. failed-but-credited)", totals["failed_clean_count"])
    row = _write_kv(ws, row, "Failed amount (Rs.)", round(totals["failed_clean_amount"], 2))
    row = _write_kv(ws, row, "Failed but credited (moves to manual reversal)", totals["failed_credited_count"])
    row = _write_kv(ws, row, "Failed-but-credited amount (Rs.)", round(totals["failed_credited_amount"], 2))
    row = _write_kv(ws, row, "Failed On-Us", onus_offus["onus"]["count"])
    row = _write_kv(ws, row, "Failed On-Us amount (Rs.)", onus_offus["onus"]["amount"])
    row = _write_kv(ws, row, "Failed Off-Us", onus_offus["offus"]["count"])
    row = _write_kv(ws, row, "Failed Off-Us amount (Rs.)", onus_offus["offus"]["amount"])
    row += 1

    row = _write_section(ws, row, "Reversal")
    row = _write_kv(ws, row, "Total reversal needed (manual + system + failed-credited)", totals["total_reversal_count"])
    row = _write_kv(ws, row, "Manual reversal (excl. already-reversed)", totals["manual_reversal_clean_count"])
    row = _write_kv(ws, row, "Manual reversal amount (Rs.)", round(totals["manual_reversal_clean_amount"], 2))
    row = _write_kv(ws, row, "System reversal (incl. already-reversed)", totals["system_reversal_clean_count"])
    row = _write_kv(ws, row, "System reversal amount (Rs.)", round(totals["system_reversal_clean_amount"], 2))
    row = _write_kv(ws, row, "Already reversed (moved from manual)", totals["already_reversed_count"])
    row = _write_kv(ws, row, "coop manual reversal", totals["coop_count"])
    row = _write_kv(ws, row, "imeremit reversal", totals["imeremit_count"])
    row = _write_kv(ws, row, "cityremit manual reversal", totals["cityremit_count"])
    row += 1

    row = _write_section(ws, row, "Data quality")
    row = _write_kv(ws, row, "Prabhu Bank reversal", totals["prabhu_rerouted"])
    row = _write_kv(ws, row, "Double-reversals prevented", totals["duplicate_skipped"])
    row = _write_kv(ws, row, "Overlapping-window duplicates skipped", totals["duplicate_source_skipped"])
    row = _write_kv(ws, row, "Unrecognized Debtor Bank rows", totals["unrecognized_debtor_bank_rows"])
    _autosize(ws)

    ws2 = wb.create_sheet("Monthly report")
    headers = [
        "Month", "Files", "Txns scanned", "Success", "Success amount (Rs.)",
        "Failed", "Failed amount (Rs.)", "Manual reversal", "System reversal",
        "Total reversal", "Timeout",
    ]
    header_row = _write_table_header(ws2, 1, headers)
    r = header_row - 1
    for mrow in monthly_qs:
        failed_credited = mrow["failed_credited_count"] or 0
        already_reversed = mrow["already_reversed_count"] or 0
        failed = max(0, (mrow["failed_total"] or 0) - failed_credited)
        reversal = max(0, (mrow["reversal_manual_kept"] or 0) - already_reversed)
        system_reversal = (mrow["reversal_system_count"] or 0) + already_reversed
        total_reversal = reversal + system_reversal + failed_credited
        month_dt = mrow["month"]
        r += 1
        values = [
            timezone.localtime(month_dt).strftime("%B %Y") if month_dt else "—",
            mrow["files"],
            mrow["total_rows"] or 0,
            mrow["success_count"] or 0,
            round(mrow["success_amount"] or 0, 2),
            failed,
            round(mrow["failed_amount"] or 0, 2),
            reversal,
            system_reversal,
            total_reversal,
            mrow["timeout_count"] or 0,
        ]
        for i, v in enumerate(values, start=1):
            ws2.cell(row=r, column=i, value=v)
    _style_data_rows(ws2, header_row, len(headers))
    _autosize(ws2)

    return _finalize_xlsx(wb, f"dashboard_summary_{suffix}.xlsx")


@login_required
def export_member_report_view(request):
    """Member-wise report (success/failed/reversal count + amount per
    Member Name) as a downloadable .xlsx — honors the page-level year/month
    filter plus this report's own From/To date range and member search."""
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    period = _period_label(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))
    suffix = _period_suffix(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))

    member_query = (request.GET.get("member_q") or "").strip()
    member_from, member_to = _parse_date_range(request, "member")

    qs = MemberAggregatorStat.objects.filter(log__in=successful_logs)
    if member_from:
        qs = qs.filter(log__created_at__date__gte=member_from)
    if member_to:
        qs = qs.filter(log__created_at__date__lte=member_to)
    qs = (
        qs.values("member_name")
        .annotate(
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_count=Sum("failed_count"),
            failed_amount=Sum("failed_amount"),
            reversal_count=Sum("reversal_count"),
            reversal_amount=Sum("reversal_amount"),
        )
        .order_by("member_name")
    )
    if member_query:
        qs = qs.filter(member_name__icontains=member_query)

    date_note = f"{member_from or 'earliest'} to {member_to or 'latest'}" if (member_from or member_to) else None

    wb, ws = _new_sheet("Member report")
    title_period = f"{period} ({date_note})" if date_note else period
    row = _write_title(ws, "Member-wise report", title_period)
    if member_query:
        row = _write_kv(ws, row, "Filtered to member matching", member_query)
        row += 1
    headers = [
        "Member", "Success", "Success amount (Rs.)", "Failed", "Failed amount (Rs.)",
        "Reversal", "Reversal amount (Rs.)",
    ]
    header_row = _write_table_header(ws, row, headers)
    r = header_row - 1
    for mrow in qs:
        r += 1
        values = [
            mrow["member_name"] or "\u2014",
            mrow["success_count"] or 0,
            round(mrow["success_amount"] or 0, 2),
            mrow["failed_count"] or 0,
            round(mrow["failed_amount"] or 0, 2),
            mrow["reversal_count"] or 0,
            round(mrow["reversal_amount"] or 0, 2),
        ]
        for i, v in enumerate(values, start=1):
            ws.cell(row=r, column=i, value=v)
    _style_data_rows(ws, header_row, len(headers))
    _autosize(ws)

    return _finalize_xlsx(wb, f"member_report_{suffix}.xlsx")


@login_required
def export_aggregator_report_view(request):
    """Aggregator-wise report (success/failed/reversal count + amount per
    Aggregator) as a downloadable .xlsx — honors the page-level year/month
    filter plus this report's own From/To date range and aggregator search."""
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    period = _period_label(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))
    suffix = _period_suffix(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))

    aggregator_query = (request.GET.get("aggregator_q") or "").strip()
    aggregator_from, aggregator_to = _parse_date_range(request, "aggregator")

    qs = MemberAggregatorStat.objects.filter(log__in=successful_logs)
    if aggregator_from:
        qs = qs.filter(log__created_at__date__gte=aggregator_from)
    if aggregator_to:
        qs = qs.filter(log__created_at__date__lte=aggregator_to)
    qs = (
        qs.values("aggregator")
        .annotate(
            success_count=Sum("success_count"),
            success_amount=Sum("success_amount"),
            failed_count=Sum("failed_count"),
            failed_amount=Sum("failed_amount"),
            reversal_count=Sum("reversal_count"),
            reversal_amount=Sum("reversal_amount"),
        )
        .order_by("aggregator")
    )
    if aggregator_query:
        qs = qs.filter(aggregator__icontains=aggregator_query)

    date_note = (
        f"{aggregator_from or 'earliest'} to {aggregator_to or 'latest'}"
        if (aggregator_from or aggregator_to) else None
    )

    wb, ws = _new_sheet("Aggregator report")
    title_period = f"{period} ({date_note})" if date_note else period
    row = _write_title(ws, "Aggregator-wise report", title_period)
    if aggregator_query:
        row = _write_kv(ws, row, "Filtered to aggregator matching", aggregator_query)
        row += 1
    headers = [
        "Aggregator", "Success", "Success amount (Rs.)", "Failed", "Failed amount (Rs.)",
        "Reversal", "Reversal amount (Rs.)",
    ]
    header_row = _write_table_header(ws, row, headers)
    r = header_row - 1
    for arow in qs:
        r += 1
        values = [
            arow["aggregator"] or "\u2014",
            arow["success_count"] or 0,
            round(arow["success_amount"] or 0, 2),
            arow["failed_count"] or 0,
            round(arow["failed_amount"] or 0, 2),
            arow["reversal_count"] or 0,
            round(arow["reversal_amount"] or 0, 2),
        ]
        for i, v in enumerate(values, start=1):
            ws.cell(row=r, column=i, value=v)
    _style_data_rows(ws, header_row, len(headers))
    _autosize(ws)

    return _finalize_xlsx(wb, f"aggregator_report_{suffix}.xlsx")


_DAY_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_DAY_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
_DAY_SUB_FONT = Font(name="Calibri", italic=True, size=9.5, color="555555")
_SUBHEAD_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_SUBHEAD_FONT = Font(name="Calibri", bold=True, size=10, color="1D4ED8")
_ROW_LABEL_FONT = Font(name="Calibri", bold=True, size=10.5)
_THIN = Side(style="thin", color="D0D0D0")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY_FMT = "#,##0.00"
_DAY_REPORT_COLS = 4  # Category/label | Count | Amount | Charge


def _write_day_subtable_header(ws, row: int, headers: list) -> int:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _SUBHEAD_FONT
        c.fill = _SUBHEAD_FILL
        c.border = _CELL_BORDER
        c.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center")
    return row + 1


def _write_day_row(ws, row: int, values: list, label_bold: bool = True) -> int:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = _CELL_BORDER
        if i == 1:
            c.font = _ROW_LABEL_FONT if label_bold else _DEFAULT_FONT
        else:
            c.font = _DEFAULT_FONT
            c.alignment = Alignment(horizontal="right")
            if isinstance(v, (int, float)) and i in (3, 4):
                c.number_format = _MONEY_FMT
    return row + 1


def _write_single_day_report(ws, row: int, detail: dict) -> int:
    """Write one day's full reconciliation report — laid out exactly like
    the website's "click a day for the full reconciliation report" modal
    (Reconciliation summary / Failure reasons / Other checks), instead of
    packing everything into one very wide row. Returns the next free row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_DAY_REPORT_COLS)
    head_cell = ws.cell(row=row, column=1, value=detail["day"])
    head_cell.font = _DAY_FONT
    head_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, _DAY_REPORT_COLS + 1):
        ws.cell(row=row, column=c).fill = _DAY_FILL
    ws.row_dimensions[row].height = 20
    row += 1

    sub_bits = [
        f"{detail['files']} file(s)",
        f"{detail['total_rows']} txns scanned",
        f"{detail['unique_new_txns']} unique new txn(s)",
    ]
    if detail.get("data_through_at"):
        sub_bits.append(f"data through {detail['data_through_at']}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_DAY_REPORT_COLS)
    sub_cell = ws.cell(row=row, column=1, value=" · ".join(sub_bits))
    sub_cell.font = _DAY_SUB_FONT
    row += 2

    # --- Reconciliation summary -----------------------------------------
    row = _write_day_subtable_header(ws, row, ["Category", "Count", "Amount (Rs.)", "Charge (Rs.)"])
    summary_rows = [
        ("Success", detail["success"]["count"], detail["success"]["amount"], None),
        ("Manual reversal", detail["manual_reversal"]["count"], detail["manual_reversal"]["amount"], detail["manual_reversal"]["charge"]),
        ("System reversal", detail["system_reversal"]["count"], detail["system_reversal"]["amount"], detail["system_reversal"]["charge"]),
        ("Failed", detail["failed"]["count"], detail["failed"]["amount"], detail["failed"]["charge"]),
        ("Failed but credited", detail["failed_credited"]["count"], detail["failed_credited"]["amount"], None),
        ("Already reversed", detail["already_reversed"]["count"], detail["already_reversed"]["amount"], None),
        ("Timeout", detail["timeout"]["count"], detail["timeout"]["amount"], None),
    ]
    for label, count, amount, charge in summary_rows:
        row = _write_day_row(ws, row, [label, count, amount, charge if charge is not None else "—"])
    row += 1

    # --- Failure reasons ---------------------------------------------------
    if detail["top_reasons"]:
        row = _write_day_subtable_header(ws, row, ["Failure reason", "Count", "% of failures", ""])
        for tr in detail["top_reasons"]:
            row = _write_day_row(ws, row, [tr["reason"], tr["count"], f"{tr['pct']}%", ""], label_bold=False)
        row += 1

    # --- Other checks --------------------------------------------------
    row = _write_day_subtable_header(ws, row, ["Other checks", "", "", "Value"])
    other_rows = [
        ("Double-reversals prevented (last-file check)", detail["duplicate_skipped"]),
        ("Overlapping-window duplicates skipped", detail["duplicate_source_skipped"]),
        ("Prabhu Bank reversal rows", detail["prabhu_rerouted"]),
        ("Unrecognized Debtor Bank rows", detail["unrecognized_debtor_bank_rows"]),
        ("Source data through", detail["data_through_at"] or "—"),
        ("Processing time (ms)", detail["processing_ms"]),
    ]
    for label, value in other_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_DAY_REPORT_COLS - 1)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _DEFAULT_FONT
        lc.border = _CELL_BORDER
        for c in range(2, _DAY_REPORT_COLS):
            ws.cell(row=row, column=c).border = _CELL_BORDER
        vc = ws.cell(row=row, column=_DAY_REPORT_COLS, value=value)
        vc.font = _DEFAULT_FONT
        vc.border = _CELL_BORDER
        vc.alignment = Alignment(horizontal="right")
        row += 1

    row += 2  # blank separator before the next day
    return row


@login_required
def export_day_breakdown_view(request):
    """Every day in the current filter (not just the paginated page shown
    on-screen), laid out as a stack of per-day report cards — same
    structure as the dashboard's "click a day for the full reconciliation
    report" modal (Reconciliation summary / Failure reasons / Other
    checks) — instead of one very wide, hard-to-read row per day.

    If ?day=YYYY-MM-DD is given (from the day modal's download button),
    the workbook is narrowed down to just that one day."""
    year_month = _apply_year_month_filter(request)
    successful_logs = year_month["logs"]
    period = _period_label(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))
    suffix = _period_suffix(year_month["selected_year"], year_month["selected_month_name"], year_month.get("selected_day"))

    single_day = (request.GET.get("day") or "").strip()
    if single_day:
        try:
            datetime.strptime(single_day, "%Y-%m-%d")
        except ValueError:
            raise Http404("Invalid day")
        successful_logs = successful_logs.filter(created_at__date=single_day)
        period = single_day
        suffix = single_day

    daily_stats, _ = _build_daily_stats(successful_logs)

    wb, ws = _new_sheet("Day by day")
    ws.cell(row=1, column=1, value="Day by day breakdown").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period}").font = _DEFAULT_FONT

    row = 4
    for day_row in daily_stats:
        row = _write_single_day_report(ws, row, day_row["detail"])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A4"

    return _finalize_xlsx(wb, f"day_by_day_breakdown_{suffix}.xlsx")